from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q, Case, When, Value, IntegerField
from django.db.models.functions import Cast
from django.views.decorators.http import require_http_methods
from django.db import transaction
from .models import Host, Idc, Cabinet, HostGroup, IpSource, SSHConfig, BastionHost, CollectTask, CollectHistory, BatchCommand, BatchCommandHistory, StaticAsset, UserProfile, Module, Role, BackupRecord, OperationLog, SparePart, SparePartType, AssetRelation, InstallHistory, LifecycleEvent, OfficePart
from django.utils import timezone
from .scheduler import update_scheduler_job
import paramiko
import io
import json
import os
import time
import threading
import logging
import uuid
import shutil
import tarfile
from datetime import datetime

# 导入硬件采集模块
from hw_collector import collect_single_server, HardwareInfo

def log_operation(user, action, target='', description='', ip_address=None):
    """记录操作日志"""
    try:
        OperationLog.objects.create(
            user=user,
            action=action,
            target=target,
            description=description,
            ip_address=ip_address
        )
    except Exception as e:
        print(f"记录操作日志失败: {e}")

def superuser_required(view_func):
    """自定义装饰器：只允许超级管理员访问"""
    def wrapped_view(request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, '您没有权限访问此页面')
            return redirect('cmdb_index')
        return view_func(request, *args, **kwargs)
    return wrapped_view

def admin_required(view_func):
    """自定义装饰器：只允许管理员用户访问（包括 is_superuser 和 is_admin）"""
    def wrapped_view(request, *args, **kwargs):
        # 检查是否是超级管理员或管理员用户
        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        
        # 检查用户是否有 is_admin 标记
        try:
            profile = request.user.userprofile
            if profile.is_admin:
                return view_func(request, *args, **kwargs)
        except UserProfile.DoesNotExist:
            pass
        
        messages.error(request, '您没有权限访问此页面')
        return redirect('cmdb_index')
    return wrapped_view

def has_admin_access(user):
    """检查用户是否有管理员权限"""
    if user.is_superuser:
        return True
    try:
        profile = user.userprofile
        return profile.is_admin
    except UserProfile.DoesNotExist:
        return False

# 解析IP范围或逗号分隔的IP列表
def parse_ip_range(ip_range):
    """解析IP范围或逗号分隔的IP列表"""
    ips = []
    parts = ip_range.split(',')
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            start_ip, end_ip = part.split('-', 1)
            start_ip = start_ip.strip()
            end_ip = end_ip.strip()

            if '.' not in end_ip:
                end_ip = '.'.join(start_ip.rsplit('.', 1)[:-1]) + '.' + end_ip

            start_parts = list(map(int, start_ip.split('.')))
            end_parts = list(map(int, end_ip.split('.')))

            if len(start_parts) == 4 and len(end_parts) == 4:
                if start_parts[0] == end_parts[0] and start_parts[1] == end_parts[1] and start_parts[2] == end_parts[2]:
                    for i in range(start_parts[3], end_parts[3] + 1):
                        ips.append(f'{start_parts[0]}.{start_parts[1]}.{start_parts[2]}.{i}')
        else:
            ips.append(part)
    return ips

# 主页视图
@login_required
def index(request):
    try:
        # 计算统计数据
        # 排除组件资产（hostname以direct-或component-开头的）
        from django.db.models import Q
        host_count = Host.objects.exclude(
            Q(hostname__startswith='direct-') | 
            Q(hostname__startswith='component-')
        ).count()
        static_asset_count = StaticAsset.objects.count()
        idc_count = Idc.objects.count()
        cabinet_count = Cabinet.objects.count()
        host_group_count = HostGroup.objects.count()
        ssh_config_count = SSHConfig.objects.count()
        collect_task_count = CollectTask.objects.count()
        batch_command_count = BatchCommand.objects.count()
        
        # 静态资产状态分布
        online_count = StaticAsset.objects.filter(status='在用').count()
        offline_count = StaticAsset.objects.filter(status='下架').count()
        other_count = static_asset_count - online_count - offline_count
        
        # 计算百分比
        online_rate = round((online_count / static_asset_count) * 100, 1) if static_asset_count > 0 else 0
        offline_rate = round((offline_count / static_asset_count) * 100, 1) if static_asset_count > 0 else 0
        other_rate = round((other_count / static_asset_count) * 100, 1) if static_asset_count > 0 else 0
        
        # 获取最近的备份记录
        recent_backups = BackupRecord.objects.order_by('-backup_time')[:10]
        
        # 获取最近的操作日志并转换格式
        operation_logs = OperationLog.objects.order_by("-created_at")[:10]
        recent_operations = []
        for op in operation_logs:
            # 构建显示内容
            action_display = op.get_action_display()
            user_display = op.user.username if op.user else '未知'
            
            # 如果操作是登录/登出，合并显示
            if op.action in ['login', 'logout']:
                content = f"{action_display} - {user_display}"
            elif op.target:
                content = f"{action_display} {op.target}"
            else:
                content = action_display
            
            # 格式化时间：YYYY-MM-DD HH:MM:SS（转换为本地时区）
            from django.utils import timezone
            local_time = timezone.localtime(op.created_at)
            time_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
            
            recent_operations.append({
                'time': time_str,
                'content': content,
                'user': user_display
            })
        
        # 静态资产类型分布
        server_type_count = StaticAsset.objects.filter(Q(server_type__icontains='机架式') | Q(server_type__icontains='服务器')).count()
        network_type_count = StaticAsset.objects.filter(server_type__icontains='网络').count()
        other_type_count = static_asset_count - server_type_count - network_type_count
        
        asset_type_dist = [
            {'name': '服务器', 'count': server_type_count, 'percentage': round((server_type_count / static_asset_count) * 100, 1) if static_asset_count > 0 else 0, 'color': '#667eea'},
            {'name': '网络设备', 'count': network_type_count, 'percentage': round((network_type_count / static_asset_count) * 100, 1) if static_asset_count > 0 else 0, 'color': '#48bb78'},
            {'name': '其他设备', 'count': other_type_count, 'percentage': round((other_type_count / static_asset_count) * 100, 1) if static_asset_count > 0 else 0, 'color': '#a0aec0'},
        ]
        
        return render(request, 'cmdb/index.html', {
            'host_count': host_count,
            'static_asset_count': static_asset_count,
            'idc_count': idc_count,
            'cabinet_count': cabinet_count,
            'host_group_count': host_group_count,
            'ssh_config_count': ssh_config_count,
            'collect_task_count': collect_task_count,
            'batch_command_count': batch_command_count,
            'recent_backups': recent_backups,
            'online_count': online_count,
            'offline_count': offline_count,
            'other_count': other_count,
            'online_rate': online_rate,
            'offline_rate': offline_rate,
            'other_rate': other_rate,
            'recent_operations': recent_operations,
            'asset_type_dist': asset_type_dist
        })
    except Exception as e:
        messages.error(request, f'系统错误: {str(e)}')
        # 出错时返回默认值
        return render(request, 'cmdb/index.html', {
            'host_count': 0,
            'static_asset_count': 0,
            'idc_count': 0,
            'cabinet_count': 0,
            'host_group_count': 0,
            'ssh_config_count': 0,
            'collect_task_count': 0,
            'batch_command_count': 0,
            'recent_collects': [],
            'online_count': 0,
            'offline_count': 0,
            'other_count': 0,
            'recent_operations': [],
            'asset_type_dist': []
        })

# 资产列表视图
@login_required
def asset_list(request):
    try:
        # 获取过滤参数
        keyword = request.GET.get('keyword', '')
        idc_id = request.GET.get('idc_id', '')
        asset_type = request.GET.get('asset_type', '')
        status = request.GET.get('status', '')
        group_filter = request.GET.get('group', '')
        cabinet_filter = request.GET.get('cabinet', '')
        
        # 构建查询（排除组件资产，组件资产hostname以direct-或component-开头）
        queryset = Host.objects.exclude(
            Q(hostname__startswith='direct-') | 
            Q(hostname__startswith='component-')
        )
        
        # 关键词搜索
        if keyword:
            queryset = queryset.filter(
                Q(hostname__icontains=keyword) |
                Q(asset_no__icontains=keyword) |
                Q(ip__icontains=keyword) |
                Q(department__icontains=keyword) |
                Q(contact_person__icontains=keyword) |
                Q(sn__icontains=keyword) |
                Q(device_model__icontains=keyword) |
                Q(cpu_model__icontains=keyword) |
                Q(os__icontains=keyword)
            )
        
        # 所属机房过滤
        if idc_id:
            queryset = queryset.filter(idc_id=idc_id)
        
        # 设备类型过滤
        if asset_type:
            queryset = queryset.filter(asset_type=asset_type)
        
        # 设备状态过滤
        if status:
            queryset = queryset.filter(status=status)
        
        # 资产组过滤
        if group_filter:
            queryset = queryset.filter(hostgroup=group_filter)
        
        # 机柜过滤
        if cabinet_filter:
            queryset = queryset.filter(cabinet_position__icontains=cabinet_filter)
        
        # 按IP地址排序
        hosts = sorted(queryset, key=lambda host: tuple(map(int, host.ip.split('.'))))
        
        # 获取下拉框数据
        idc_list = Idc.objects.all()
        host_groups = HostGroup.objects.all()
        
        # 获取机柜列表
        try:
            cabinet_list = list(Host.objects.values_list('cabinet_position', flat=True).distinct().order_by('cabinet_position'))
            cabinet_list = [c for c in cabinet_list if c]
        except:
            cabinet_list = []
        
        # 设备类型选项
        asset_types = [
            ('1', '物理机'),
            ('2', '虚拟机'),
            ('3', '容器'),
            ('4', '网络设备'),
            ('5', '安全设备'),
            ('6', '其他')
        ]
        
        # 设备状态选项
        asset_status = [
            ('1', '在线'),
            ('2', '离线'),
            ('3', '故障'),
            ('4', '维护')
        ]
        
        # 计算统计数据
        total_count = Host.objects.exclude(
            Q(hostname__startswith='direct-') | 
            Q(hostname__startswith='component-')
        ).count()
        filtered_count = len(hosts)
        has_filter = keyword or idc_id or asset_type or status or group_filter or cabinet_filter
        
        response = render(request, 'cmdb/asset_list.html', {
            'hosts': hosts,
            'keyword': keyword,
            'idc_id': idc_id,
            'asset_type': asset_type,
            'status': status,
            'group_filter': group_filter,
            'cabinet_filter': cabinet_filter,
            'idc_list': idc_list,
            'host_groups': host_groups,
            'cabinet_list': cabinet_list,
            'asset_types': asset_types,
            'asset_status': asset_status,
            'total_count': total_count,
            'filtered_count': filtered_count,
            'has_filter': has_filter
        })
        
        # 添加缓存控制头，防止浏览器缓存页面
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        
        return response
    except Exception as e:
        messages.error(request, f'获取资产列表失败: {str(e)}')
        return render(request, 'cmdb/asset_list.html', {'hosts': []})

# 资产相关视图
@login_required
def asset_add(request):
    try:
        # 简化实现
        return render(request, 'cmdb/asset_form.html', {'action': 'add'})
    except Exception as e:
        messages.error(request, f'添加资产失败: {str(e)}')
        return redirect('asset_list')

@login_required
def asset_detail(request, asset_id):
    try:
        host = get_object_or_404(Host, id=asset_id)
        return render(request, 'cmdb/asset_detail.html', {'host': host})
    except Exception as e:
        messages.error(request, f'获取资产详情失败: {str(e)}')
        return redirect('asset_list')

@login_required
def asset_edit(request, asset_id):
    try:
        host = get_object_or_404(Host, id=asset_id)
        if request.method == 'POST':
            try:
                host.hostname = request.POST.get('hostname')
                host.asset_no = request.POST.get('asset_no')
                host.cabinet_position = request.POST.get('cabinet_position')
                host.department = request.POST.get('department')
                host.asset_type = request.POST.get('asset_type')
                host.ip = request.POST.get('ip')
                host.contact_person = request.POST.get('contact_person')
                host.device_model = request.POST.get('device_model')
                host.cpu_model = request.POST.get('cpu_model')
                host.cpu_num = request.POST.get('cpu_num')
                host.cpu_cores = request.POST.get('cpu_cores')
                host.gpu_model = request.POST.get('gpu_model')
                host.memory = request.POST.get('memory')
                host.disk = request.POST.get('disk')
                host.os = request.POST.get('os')
                host.sn = request.POST.get('sn')
                host.bm_ip = request.POST.get('bm_ip')
                host.up_time = request.POST.get('up_time')
                host.status = request.POST.get('status')
                host.memo = request.POST.get('memo')
                host.idc_id = request.POST.get('idc')
                host.cabinet_id = request.POST.get('cabinet')
                host.hostgroup = request.POST.get('group')
                host.save()
                messages.success(request, f'资产 {host.hostname} 更新成功')
                return redirect('asset_list')
            except Exception as e:
                messages.error(request, f'更新失败: {str(e)}')
        return render(request, 'cmdb/asset_form.html', {'host': host, 'action': 'edit'})
    except Exception as e:
        messages.error(request, f'编辑资产失败: {str(e)}')
        return redirect('asset_list')

@login_required
def asset_delete(request, asset_id):
    try:
        asset = get_object_or_404(Host, id=asset_id)
        asset_name = asset.hostname
        asset.delete()

        log_operation(request.user, 'delete', f'动态资产: {asset_name}', f'删除动态资产: {asset_name}', request.META.get('REMOTE_ADDR'))

        messages.success(request, f'资产 {asset_name} 删除成功')
        return redirect('asset_list')
    except Exception as e:
        messages.error(request, f'删除资产失败: {str(e)}')
        return redirect('asset_list')

@login_required
def import_assets_excel(request):
    try:
        # 简化实现
        return render(request, 'cmdb/asset_import.html')
    except Exception as e:
        messages.error(request, f'导入资产失败: {str(e)}')
        return redirect('asset_list')

@login_required
def export_assets_excel(request):
    try:
        import io
        from openpyxl import Workbook
        from cmdb.models import ASSET_STATUS, ASSET_TYPE
        from datetime import datetime
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '资产列表'
        
        # 添加表头（与页面展示顺序一致）
        headers = ['主机名', '资产编号', '机柜位置', '部门/团队', '设备类型', 'IP地址', '责任人', '品牌型号', 'CPU型号', 'CPU数量', 'CPU核数', 'GPU', '内存', '硬盘', '操作系统', '序列号', '带外IP', '上架时间', '设备状态', '所属机房', '备注']
        ws.append(headers)
        
        # 获取资产数据（排除组件资产）
        hosts = Host.objects.exclude(
            Q(hostname__startswith='direct-') | 
            Q(hostname__startswith='component-')
        )
        for host in hosts:
            row = [
                host.hostname,
                host.asset_no or '',
                host.cabinet_position or '',
                host.department or '',
                dict(ASSET_TYPE).get(host.asset_type, ''),
                host.ip,
                host.contact_person or '',
                host.device_model or '',
                host.cpu_model or '',
                host.cpu_num or '',
                host.cpu_cores or '',
                host.gpu_model or '',
                host.memory or '',
                host.disk or '',
                host.os or '',
                host.sn or '',
                host.bm_ip or '',
                host.up_time or '',
                dict(ASSET_STATUS).get(host.status, ''),
                host.idc.name if host.idc else '',
                host.memo or ''
            ]
            ws.append(row)
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f'assets_export_{timestamp}.xlsx'
        
        # 设置响应头
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'导出失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})

@login_required
def export_assets_csv(request):
    try:
        import csv
        import io
        from cmdb.models import ASSET_STATUS, ASSET_TYPE
        from datetime import datetime
        
        # 创建内存文件
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 添加表头（与页面展示顺序一致）
        headers = ['主机名', '资产编号', '机柜位置', '部门/团队', '设备类型', 'IP地址', '责任人', '品牌型号', 'CPU型号', 'CPU数量', 'CPU核数', 'GPU', '内存', '硬盘', '操作系统', '序列号', '带外IP', '上架时间', '设备状态', '所属机房', '备注']
        writer.writerow(headers)
        
        # 获取资产数据（排除组件资产）
        hosts = Host.objects.exclude(
            Q(hostname__startswith='direct-') | 
            Q(hostname__startswith='component-')
        )
        for host in hosts:
            row = [
                host.hostname,
                host.asset_no or '',
                host.cabinet_position or '',
                host.department or '',
                dict(ASSET_TYPE).get(host.asset_type, ''),
                host.ip,
                host.contact_person or '',
                host.device_model or '',
                host.cpu_model or '',
                host.cpu_num or '',
                host.cpu_cores or '',
                host.gpu_model or '',
                host.memory or '',
                host.disk or '',
                host.os or '',
                host.sn or '',
                host.bm_ip or '',
                host.up_time or '',
                dict(ASSET_STATUS).get(host.status, ''),
                host.idc.name if host.idc else '',
                host.memo or ''
            ]
            writer.writerow(row)
        
        output.seek(0)
        
        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f'assets_export_{timestamp}.csv'
        
        # 设置响应头
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'导出失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})

# 批量删除资产
@login_required
def asset_batch_delete(request):
    try:
        if request.method == 'POST':
            import json
            data = json.loads(request.body)
            asset_ids = data.get('ids', [])

            if not asset_ids:
                return JsonResponse({'success': False, 'message': '请选择要删除的资产'})

            # 删除选定的资产
            deleted_count = Host.objects.filter(id__in=asset_ids).delete()[0]

            log_operation(request.user, 'delete', '动态资产', f'批量删除 {deleted_count} 个动态资产', request.META.get('REMOTE_ADDR'))

            return JsonResponse({'success': True, 'message': f'成功删除 {deleted_count} 个资产'})
        else:
            return JsonResponse({'success': False, 'message': '只支持POST请求'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'批量删除失败: {str(e)}'})

@login_required
def api_hosts(request):
    try:
        # 简化实现
        return JsonResponse({'success': True, 'hosts': []})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'获取主机列表失败: {str(e)}'})

# IDC相关视图
@login_required
def idc_list(request):
    try:
        idcs = Idc.objects.all()
        return render(request, 'cmdb/idc_list.html', {'idcs': idcs})
    except Exception as e:
        messages.error(request, f'获取IDC列表失败: {str(e)}')
        return render(request, 'cmdb/idc_list.html', {'idcs': []})

@login_required
def idc_add(request):
    try:
        if request.method == 'POST':
            # 处理表单提交
            ids = request.POST.get('ids')
            name = request.POST.get('name')
            address = request.POST.get('address')
            tel = request.POST.get('phone', '') or ''
            contact = request.POST.get('contact', '') or ''
            memo = request.POST.get('desc', '') or ''
            
            # 创建IDC
            idc = Idc(ids=ids, name=name, address=address, tel=tel, contact=contact, memo=memo)
            idc.save()

            log_operation(request.user, 'add', f'IDC: {name}', f'添加IDC机房: {name}', request.META.get('REMOTE_ADDR'))

            messages.success(request, f'IDC {name} 添加成功')
            return redirect('idc_list')
        else:
            # 渲染表单
            return render(request, 'cmdb/idc_form.html', {'action': 'add'})
    except Exception as e:
        messages.error(request, f'添加IDC失败: {str(e)}')
        return redirect('idc_list')

@login_required
def idc_edit(request, idc_id):
    try:
        idc = get_object_or_404(Idc, id=idc_id)
        
        if request.method == 'POST':
            # 处理表单提交
            ids = request.POST.get('ids')
            name = request.POST.get('name')
            address = request.POST.get('address')
            tel = request.POST.get('phone', '') or ''
            contact = request.POST.get('contact', '') or ''
            memo = request.POST.get('desc', '') or ''
            
            # 更新IDC信息
            idc.ids = ids
            idc.name = name
            idc.address = address
            idc.tel = tel
            idc.contact = contact
            idc.memo = memo
            idc.save()
            
            messages.success(request, f'IDC {name} 更新成功')
            return redirect('idc_list')
        else:
            # 渲染表单
            return render(request, 'cmdb/idc_form.html', {'idc': idc, 'action': 'edit'})
    except Exception as e:
        messages.error(request, f'编辑IDC失败: {str(e)}')
        return redirect('idc_list')

@login_required
def idc_delete(request, idc_id):
    try:
        idc = get_object_or_404(Idc, id=idc_id)
        idc_name = idc.name
        idc.delete()

        log_operation(request.user, 'delete', f'IDC: {idc_name}', f'删除IDC机房: {idc_name}', request.META.get('REMOTE_ADDR'))

        messages.success(request, f'IDC {idc_name} 删除成功')
        return redirect('idc_list')
    except Exception as e:
        messages.error(request, f'删除IDC失败: {str(e)}')
        return redirect('idc_list')

# 机柜相关视图
@login_required
def cabinet_list(request):
    try:
        idc_id = request.GET.get('idc_id', '')
        cabinets = Cabinet.objects.all()
        
        if idc_id:
            cabinets = cabinets.filter(idc_id=idc_id)
        
        idc_list = Idc.objects.all()
        
        response = render(request, 'cmdb/cabinet_list.html', {
            'cabinets': cabinets,
            'idc_list': idc_list,
            'selected_idc_id': idc_id
        })
        
        # 添加缓存控制头，防止浏览器缓存页面
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        
        return response
    except Exception as e:
        messages.error(request, f'获取机柜列表失败: {str(e)}')
        return render(request, 'cmdb/cabinet_list.html', {'cabinets': []})

@login_required
def cabinet_add(request):
    try:
        if request.method == 'POST':
            # 处理表单提交
            name = request.POST.get('name')
            idc_id = request.POST.get('idc')
            desc = request.POST.get('desc')
            
            # 创建机柜
            cabinet = Cabinet(name=name, idc_id=idc_id, desc=desc)
            cabinet.save()

            log_operation(request.user, 'add', f'机柜: {name}', f'添加机柜: {name}', request.META.get('REMOTE_ADDR'))

            messages.success(request, f'机柜 {name} 添加成功')
            return redirect('cabinet_list')
        else:
            # 渲染表单
            idc_list = Idc.objects.all()
            return render(request, 'cmdb/cabinet_form.html', {'action': 'add', 'idc_list': idc_list})
    except Exception as e:
        messages.error(request, f'添加机柜失败: {str(e)}')
        return redirect('cabinet_list')

@login_required
def cabinet_edit(request, cabinet_id):
    try:
        cabinet = get_object_or_404(Cabinet, id=cabinet_id)
        
        if request.method == 'POST':
            # 处理表单提交
            name = request.POST.get('name')
            idc_id = request.POST.get('idc')
            desc = request.POST.get('desc')
            
            # 更新机柜信息
            cabinet.name = name
            cabinet.idc_id = idc_id
            cabinet.desc = desc
            cabinet.save()
            
            messages.success(request, f'机柜 {name} 更新成功')
            return redirect('cabinet_list')
        else:
            # 渲染表单
            idc_list = Idc.objects.all()
            return render(request, 'cmdb/cabinet_form.html', {'cabinet': cabinet, 'action': 'edit', 'idc_list': idc_list})
    except Exception as e:
        messages.error(request, f'编辑机柜失败: {str(e)}')
        return redirect('cabinet_list')

@login_required
def cabinet_delete(request, cabinet_id):
    try:
        cabinet = get_object_or_404(Cabinet, id=cabinet_id)
        cabinet_name = cabinet.name
        cabinet.delete()

        log_operation(request.user, 'delete', f'机柜: {cabinet_name}', f'删除机柜: {cabinet_name}', request.META.get('REMOTE_ADDR'))

        messages.success(request, f'机柜 {cabinet_name} 删除成功')
        return redirect('cabinet_list')
    except Exception as e:
        messages.error(request, f'删除机柜失败: {str(e)}')
        return redirect('cabinet_list')

# 主机组相关视图
@login_required
def group_list(request):
    try:
        groups = HostGroup.objects.all()
        return render(request, 'cmdb/group_list.html', {'groups': groups})
    except Exception as e:
        messages.error(request, f'获取主机组列表失败: {str(e)}')
        return render(request, 'cmdb/group_list.html', {'groups': []})

@login_required
def group_add(request):
    try:
        if request.method == 'POST':
            # 处理表单提交
            name = request.POST.get('name')
            desc = request.POST.get('desc')
            server_ids = request.POST.getlist('servers')
            
            # 创建资产组
            group = HostGroup(name=name, desc=desc)
            group.save()
            
            # 添加选中的服务器
            if server_ids:
                servers = Host.objects.filter(id__in=server_ids)
                group.serverList.add(*servers)
            
            messages.success(request, f'资产组 {name} 添加成功')
            return redirect('group_list')
        else:
            # 渲染表单
            # 获取所有主机作为可用服务器（排除组件资产）
            available_hosts = Host.objects.exclude(
                Q(hostname__startswith='direct-') | 
                Q(hostname__startswith='component-')
            )
            # 按主机名排序
            available_hosts = sorted(available_hosts, key=lambda host: host.hostname)
            # 新组没有已选服务器
            selected_hosts = []
            return render(request, 'cmdb/group_form.html', {
                'action': 'add',
                'available_hosts': available_hosts,
                'selected_hosts': selected_hosts
            })
    except Exception as e:
        messages.error(request, f'添加主机组失败: {str(e)}')
        return redirect('group_list')

@login_required
def group_edit(request, group_id):
    try:
        group = get_object_or_404(HostGroup, id=group_id)
        
        if request.method == 'POST':
            # 处理表单提交
            name = request.POST.get('name')
            desc = request.POST.get('desc')
            server_ids = request.POST.getlist('servers')
            
            # 更新资产组信息
            group.name = name
            group.desc = desc
            group.save()

            log_operation(request.user, 'edit', f'资产组: {name}', f'更新资产组: {name}', request.META.get('REMOTE_ADDR'))

            messages.success(request, f'资产组 {name} 更新成功')
            return redirect('group_list')
        else:
            # 渲染表单
            # 获取所有主机（排除组件资产）
            all_hosts = Host.objects.exclude(
                Q(hostname__startswith='direct-') | 
                Q(hostname__startswith='component-')
            )
            # 按主机名排序
            all_hosts = sorted(all_hosts, key=lambda host: host.hostname)
            # 已选服务器
            selected_hosts = list(group.serverList.all())
            # 可用服务器（未被选中的）
            available_hosts = [host for host in all_hosts if host not in selected_hosts]
            return render(request, 'cmdb/group_form.html', {
                'group': group,
                'action': 'edit',
                'available_hosts': available_hosts,
                'selected_hosts': selected_hosts
            })
    except Exception as e:
        messages.error(request, f'编辑主机组失败: {str(e)}')
        return redirect('group_list')

@login_required
def group_delete(request, group_id):
    try:
        group = get_object_or_404(HostGroup, id=group_id)
        group.delete()
        messages.success(request, f'主机组 {group.name} 删除成功')
        return redirect('group_list')
    except Exception as e:
        messages.error(request, f'删除主机组失败: {str(e)}')
        return redirect('group_list')

# 目标服务器相关视图
@login_required
def target_server_list(request):
    try:
        configs = SSHConfig.objects.all()
        return render(request, 'cmdb/target_server_list.html', {'configs': configs})
    except Exception as e:
        messages.error(request, f'获取目标服务器列表失败: {str(e)}')
        return render(request, 'cmdb/target_server_list.html', {'configs': []})

@login_required
def target_server_add(request):
    try:
        if request.method == 'POST':
            config = SSHConfig()
            config.name = request.POST.get('name')
            config.host = request.POST.get('host')
            config.port = int(request.POST.get('port', 22))
            config.username = request.POST.get('username')
            config.password = request.POST.get('password')
            config.private_key = request.POST.get('private_key')
            config.collect_asset_types = request.POST.get('collect_asset_types', '1')
            config.is_enabled = request.POST.get('is_enabled') == 'on'
            config.memo = request.POST.get('memo')
            config.save()
            messages.success(request, f'目标服务器 {config.name} 添加成功')
            return redirect('target_server_list')
        return render(request, 'cmdb/target_server_form.html', {'action': 'add'})
    except Exception as e:
        messages.error(request, f'添加目标服务器失败: {str(e)}')
        return redirect('target_server_list')

@login_required
def target_server_edit(request, server_id):
    try:
        config = get_object_or_404(SSHConfig, id=server_id)
        if request.method == 'POST':
            config.name = request.POST.get('name')
            config.host = request.POST.get('host')
            config.port = int(request.POST.get('port', 22))
            config.username = request.POST.get('username')
            config.password = request.POST.get('password')
            config.private_key = request.POST.get('private_key')
            config.collect_asset_types = request.POST.get('collect_asset_types', '1')
            config.is_enabled = request.POST.get('is_enabled') == 'on'
            config.memo = request.POST.get('memo')
            config.save()
            messages.success(request, f'目标服务器 {config.name} 编辑成功')
            return redirect('target_server_list')
        return render(request, 'cmdb/target_server_form.html', {'action': 'edit', 'config': config})
    except Exception as e:
        messages.error(request, f'编辑目标服务器失败: {str(e)}')
        return redirect('target_server_list')

@login_required
def target_server_delete(request, server_id):
    try:
        config = get_object_or_404(SSHConfig, id=server_id)
        config_name = config.name
        config.delete()
        messages.success(request, f'目标服务器 {config_name} 删除成功')
        return redirect('target_server_list')
    except Exception as e:
        messages.error(request, f'删除目标服务器失败: {str(e)}')
        return redirect('target_server_list')

@login_required
def target_server_batch_delete(request):
    try:
        return JsonResponse({'success': True, 'message': '批量删除成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'批量删除失败: {str(e)}'})

@login_required
def export_target_server_excel(request):
    try:
        import io
        from openpyxl import Workbook
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = '目标服务器列表'

        headers = ['配置名称', 'SSH服务器IP', 'SSH端口', '用户名', '密码', '私钥内容', '采集资产类型', '启用状态', '备注', '创建时间', '更新时间']
        ws.append(headers)

        configs = SSHConfig.objects.all()
        for config in configs:
            collect_types = []
            for t in config.collect_asset_types.split(','):
                t = t.strip()
                if t == '1':
                    collect_types.append('物理机')
                elif t == '2':
                    collect_types.append('虚拟机')
                elif t == '3':
                    collect_types.append('容器')
                elif t == '6':
                    collect_types.append('其他')
            collect_types_str = ','.join(collect_types) if collect_types else ''

            row = [
                config.name,
                config.host,
                config.port,
                config.username,
                config.password or '',
                config.private_key or '',
                collect_types_str,
                '启用' if config.is_enabled else '禁用',
                config.memo or '',
                config.created_at.strftime('%Y-%m-%d %H:%M:%S') if config.created_at else '',
                config.updated_at.strftime('%Y-%m-%d %H:%M:%S') if config.updated_at else ''
            ]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f'target_servers_export_{timestamp}.xlsx'

        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'导出失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})

@login_required
def import_target_server_excel(request):
    try:
        return render(request, 'cmdb/target_server_import.html')
    except Exception as e:
        messages.error(request, f'导入目标服务器失败: {str(e)}')
        return redirect('target_server_list')

# 跳板机相关视图
@login_required
@admin_required
def bastion_list(request):
    try:
        bastions = BastionHost.objects.all()
        return render(request, 'cmdb/bastion_list.html', {'bastions': bastions})
    except Exception as e:
        messages.error(request, f'获取跳板机列表失败: {str(e)}')
        return render(request, 'cmdb/bastion_list.html', {'bastions': []})

@login_required
def bastion_add(request):
    try:
        if request.method == 'POST':
            # 处理表单提交
            bastion = BastionHost(
                name=request.POST.get('name'),
                host=request.POST.get('host'),
                port=int(request.POST.get('port', 22)),
                username=request.POST.get('username'),
                password=request.POST.get('password'),
                private_key=request.POST.get('private_key'),
                is_enabled=request.POST.get('is_enabled') == 'on',
                memo=request.POST.get('memo')
            )
            bastion.save()
            messages.success(request, f'跳板机 {bastion.name} 添加成功')
            return redirect('bastion_list')
        return render(request, 'cmdb/bastion_form.html', {'action': 'add'})
    except Exception as e:
        messages.error(request, f'添加跳板机失败: {str(e)}')
        return redirect('bastion_list')

@login_required
def bastion_edit(request, bastion_id):
    try:
        bastion = get_object_or_404(BastionHost, id=bastion_id)
        if request.method == 'POST':
            # 处理表单提交
            bastion.name = request.POST.get('name')
            bastion.host = request.POST.get('host')
            bastion.port = int(request.POST.get('port', 22))
            bastion.username = request.POST.get('username')
            bastion.password = request.POST.get('password')
            bastion.private_key = request.POST.get('private_key')
            bastion.is_enabled = request.POST.get('is_enabled') == 'on'
            bastion.memo = request.POST.get('memo')
            bastion.save()
            messages.success(request, f'跳板机 {bastion.name} 更新成功')
            return redirect('bastion_list')
        return render(request, 'cmdb/bastion_form.html', {'bastion': bastion, 'action': 'edit'})
    except Exception as e:
        messages.error(request, f'编辑跳板机失败: {str(e)}')
        return redirect('bastion_list')

@login_required
def bastion_delete(request, bastion_id):
    try:
        bastion = get_object_or_404(BastionHost, id=bastion_id)
        bastion.delete()
        messages.success(request, f'跳板机 {bastion.name} 删除成功')
        return redirect('bastion_list')
    except Exception as e:
        messages.error(request, f'删除跳板机失败: {str(e)}')
        return redirect('bastion_list')

# 采集任务相关视图
@login_required
def collect_task_list(request):
    try:
        tasks = CollectTask.objects.all()
        return render(request, 'cmdb/collect_task_list.html', {'tasks': tasks})
    except Exception as e:
        messages.error(request, f'获取采集任务列表失败: {str(e)}')
        return render(request, 'cmdb/collect_task_list.html', {'tasks': []})

@login_required
def collect_task_add(request):
    try:
        bastions = BastionHost.objects.filter(is_enabled=True)
        if request.method == 'POST':
            # 处理表单提交
            bastion_id = request.POST.get('bastion')
            bastion = BastionHost.objects.get(id=bastion_id) if bastion_id else None
            
            task = CollectTask(
                name=request.POST.get('name'),
                bastion=bastion,
                target_hosts=request.POST.get('target_hosts'),
                collect_online_only=request.POST.get('collect_online_only') == 'on',
                jump_via_bastion=request.POST.get('jump_via_bastion') == 'on',
                is_enabled=request.POST.get('is_enabled') == 'on',
                update_hostname=request.POST.get('update_hostname') == 'on',
                update_os=request.POST.get('update_os') == 'on',
                update_cpu=request.POST.get('update_cpu') == 'on',
                update_memory=request.POST.get('update_memory') == 'on',
                update_disk=request.POST.get('update_disk') == 'on',
                update_gpu=request.POST.get('update_gpu') == 'on',
                update_device_info=request.POST.get('update_device_info') == 'on',
                update_sn=request.POST.get('update_sn') == 'on',
                is_auto_collect=request.POST.get('is_auto_collect') == 'on',
                cron_expression=request.POST.get('cron_expression', '0 2 * * *'),
                memo=request.POST.get('memo')
            )
            task.save()
            
            # 如果启用了定时任务，更新调度器
            if task.is_auto_collect:
                update_scheduler_job()
            
            messages.success(request, f'采集任务 {task.name} 创建成功')
            return redirect('collect_task_list')
        return render(request, 'cmdb/collect_task_form.html', {'action': 'add', 'bastions': bastions})
    except Exception as e:
        messages.error(request, f'添加采集任务失败: {str(e)}')
        return redirect('collect_task_list')

@login_required
def collect_task_edit(request, task_id):
    try:
        task = get_object_or_404(CollectTask, id=task_id)
        bastions = BastionHost.objects.filter(is_enabled=True)
        if request.method == 'POST':
            # 处理表单提交
            bastion_id = request.POST.get('bastion')
            bastion = BastionHost.objects.get(id=bastion_id) if bastion_id else None
            
            task.name = request.POST.get('name')
            task.bastion = bastion
            task.target_hosts = request.POST.get('target_hosts')
            task.collect_online_only = request.POST.get('collect_online_only') == 'on'
            task.jump_via_bastion = request.POST.get('jump_via_bastion') == 'on'
            task.is_enabled = request.POST.get('is_enabled') == 'on'
            task.update_hostname = request.POST.get('update_hostname') == 'on'
            task.update_os = request.POST.get('update_os') == 'on'
            task.update_cpu = request.POST.get('update_cpu') == 'on'
            task.update_memory = request.POST.get('update_memory') == 'on'
            task.update_disk = request.POST.get('update_disk') == 'on'
            task.update_gpu = request.POST.get('update_gpu') == 'on'
            task.update_device_info = request.POST.get('update_device_info') == 'on'
            task.update_sn = request.POST.get('update_sn') == 'on'
            task.is_auto_collect = request.POST.get('is_auto_collect') == 'on'
            task.cron_expression = request.POST.get('cron_expression', '0 2 * * *')
            task.memo = request.POST.get('memo')
            task.save()
            
            # 更新调度器
            update_scheduler_job()
            
            messages.success(request, f'采集任务 {task.name} 更新成功')
            return redirect('collect_task_list')
        return render(request, 'cmdb/collect_task_form.html', {'task': task, 'action': 'edit', 'bastions': bastions})
    except Exception as e:
        messages.error(request, f'编辑采集任务失败: {str(e)}')
        return redirect('collect_task_list')

@login_required
def collect_task_delete(request, task_id):
    try:
        task = get_object_or_404(CollectTask, id=task_id)
        task.delete()
        messages.success(request, f'采集任务 {task.name} 删除成功')
        return redirect('collect_task_list')
    except Exception as e:
        messages.error(request, f'删除采集任务失败: {str(e)}')
        return redirect('collect_task_list')

@login_required
def collect_task_progress(request, progress_file):
    import os
    import json
    
    try:
        temp_dir = '/tmp'
        progress_path = os.path.join(temp_dir, progress_file)
        
        if os.path.exists(progress_path):
            with open(progress_path, 'r') as f:
                progress_data = json.load(f)
            return JsonResponse(progress_data)
        else:
            return JsonResponse({'error': 'Progress file not found'})
    except Exception as e:
        return JsonResponse({'error': str(e)})

@login_required
def collect_history(request):
    try:
        histories = CollectHistory.objects.all().order_by('-collect_time')
        return render(request, 'cmdb/collect_history.html', {'histories': histories})
    except Exception as e:
        messages.error(request, f'获取采集历史失败: {str(e)}')
        return render(request, 'cmdb/collect_history.html', {'histories': []})

@login_required
def run_collect_task(request, task_id):
    import logging
    import time
    import threading
    from django.shortcuts import render
    logger = logging.getLogger(__name__)

    try:
        logger.error(f"[CollectTask] 开始执行采集任务 {task_id}")

        if not CollectTask or not CollectHistory:
            raise Exception('模型未导入')

        task = get_object_or_404(CollectTask, id=task_id)
        logger.error(f"[CollectTask] 找到采集任务: {task.name}, 目标主机: {task.target_hosts}")

        progress_file = f'collect_task_progress_{task_id}_{int(time.time())}.json'
        logger.error(f"[CollectTask] 进度文件: {progress_file}")

        def execute_collection():
            import json
            import os
            from datetime import datetime

            logger.error(f"[CollectTask] execute_collection 开始执行")

            progress_data = {
                'task_id': task_id,
                'total': 0,
                'completed': 0,
                'status': 'running',
                'stopped': False,
                'finished': False,
                'results': []
            }

            temp_dir = '/tmp'
            progress_path = os.path.join(temp_dir, progress_file)

            # 保存初始进度
            with open(progress_path, 'w') as f:
                json.dump(progress_data, f)

            logger.error(f"[CollectTask] 进度文件已创建: {progress_path}")

            # 解析目标主机
            target_ips = []
            if task.target_group:
                # 从资产组获取主机
                hosts = task.target_group.serverList.all()
                target_ips.extend([host.ip for host in hosts])
                logger.error(f"[CollectTask] 从资产组获取主机: {[host.ip for host in hosts]}")

            if task.target_hosts:
                # 解析IP范围或逗号分隔的IP列表
                parsed_ips = parse_ip_range(task.target_hosts)
                logger.error(f"[CollectTask] 解析目标主机: {task.target_hosts} -> {parsed_ips}")
                target_ips.extend(parsed_ips)

            # 去重并按IP地址排序
            target_ips = list(set(target_ips))
            # 按IP地址排序，确保从192.168.12.1开始
            target_ips.sort(key=lambda ip: tuple(map(int, ip.split('.'))))
            progress_data['total'] = len(target_ips)

            # 保存更新的进度
            with open(progress_path, 'w') as f:
                json.dump(progress_data, f)

            # 执行采集
            for ip in target_ips:
                # 检查是否停止
                try:
                    with open(progress_path, 'r') as f:
                        progress_data = json.load(f)
                    if progress_data.get('stopped', False):
                        # 标记任务为已停止
                        progress_data['stopped'] = True
                        progress_data['status'] = 'stopped'
                        progress_data['finished'] = True
                        with open(progress_path, 'w') as f:
                            json.dump(progress_data, f)
                        logger.error(f"[CollectTask] 任务已停止")
                        break
                except:
                    pass

                result = {
                    'ip': ip,
                    'hostname': '',
                    'status': 'failed',
                    'output': '',
                    'error': ''
                }

                try:
                    # 检查是否停止
                    try:
                        with open(progress_path, 'r') as f:
                            progress_data = json.load(f)
                        if progress_data.get('stopped', False):
                            # 标记任务为已停止
                            progress_data['stopped'] = True
                            progress_data['status'] = 'stopped'
                            progress_data['finished'] = True
                            with open(progress_path, 'w') as f:
                                json.dump(progress_data, f)
                            logger.error(f"[CollectTask] 任务已停止")
                            break
                    except:
                        pass

                    # 执行采集（使用真实的硬件采集工具）
                    import socket
                    
                    # 检测目标主机是否可访问
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(2)  # 2秒超时
                    result_port = sock.connect_ex((ip, 22))  # 尝试连接SSH端口
                    sock.close()
                    
                    if result_port != 0:
                        # 主机不可访问，标记为失败
                        result['error'] = f'主机 {ip} 不可访问（SSH端口22未开放）'
                        result['status'] = 'failed'
                        
                        # 更新设备状态为其他
                        try:
                            host = Host.objects.get(ip=ip)
                            host.status = '4'  # 其他
                            host.save()
                            logger.error(f"[CollectTask] 资产 {ip} 不可访问，状态更新为其他")
                        except Host.DoesNotExist:
                            # 资产不存在，不需要处理
                            pass
                        except Exception as e:
                            logger.error(f"[CollectTask] 更新资产 {ip} 状态失败: {str(e)}")
                    else:
                        # 再次检查是否停止
                        try:
                            with open(progress_path, 'r') as f:
                                progress_data = json.load(f)
                            if progress_data.get('stopped', False):
                                # 标记任务为已停止
                                progress_data['stopped'] = True
                                progress_data['status'] = 'stopped'
                                progress_data['finished'] = True
                                with open(progress_path, 'w') as f:
                                    json.dump(progress_data, f)
                                logger.error(f"[CollectTask] 任务已停止")
                                break
                        except:
                            pass

                        # 主机可访问，使用真实的硬件采集工具
                        # 查找可用的SSH配置
                        ssh_config = SSHConfig.objects.filter(is_enabled=True).first()
                        if not ssh_config:
                            result['error'] = '没有可用的SSH配置'
                            result['status'] = 'failed'
                        else:
                            # 再次检查是否停止
                            try:
                                with open(progress_path, 'r') as f:
                                    progress_data = json.load(f)
                                if progress_data.get('stopped', False):
                                    # 标记任务为已停止
                                    progress_data['stopped'] = True
                                    progress_data['status'] = 'stopped'
                                    progress_data['finished'] = True
                                    with open(progress_path, 'w') as f:
                                        json.dump(progress_data, f)
                                    logger.error(f"[CollectTask] 任务已停止")
                                    break
                            except:
                                pass

                            # 执行真实的硬件采集（使用超时控制）
                            logger.error(f"[CollectTask] 开始采集主机 {ip}")
                            
                            # 使用线程执行采集，以便可以中断
                            import threading
                            import time
                            hardware_info = None
                            采集完成 = False
                            
                            def 采集_thread():
                                nonlocal hardware_info, 采集完成
                                try:
                                    hardware_info = collect_single_server(
                                        ip, 
                                        ssh_config.port, 
                                        ssh_config.username, 
                                        ssh_config.password,
                                        collect_disk=task.update_disk
                                    )
                                finally:
                                    采集完成 = True
                            
                            # 启动采集线程
                            thread = threading.Thread(target=采集_thread)
                            thread.daemon = True
                            thread.start()
                            
                            # 等待采集完成，同时定期检查是否停止
                            start_time = time.time()
                            timeout = 30  # 30秒超时
                            while not 采集完成 and time.time() - start_time < timeout:
                                # 检查是否停止
                                try:
                                    with open(progress_path, 'r') as f:
                                        progress_data = json.load(f)
                                    if progress_data.get('stopped', False):
                                        # 标记为停止状态
                                        result['error'] = '任务已停止'
                                        result['status'] = 'stopped'
                                        # 标记任务为已停止
                                        progress_data['stopped'] = True
                                        progress_data['status'] = 'stopped'
                                        progress_data['finished'] = True
                                        with open(progress_path, 'w') as f:
                                            json.dump(progress_data, f)
                                        logger.error(f"[CollectTask] 任务已停止")
                                        break
                                except:
                                    pass
                                time.sleep(1)  # 1秒检查一次
                            
                            if not 采集完成:
                                # 采集超时
                                result['error'] = '采集超时（30秒）'
                                result['status'] = 'failed'
                            else:
                                if hardware_info.error:
                                    result['error'] = hardware_info.error
                                    result['status'] = 'failed'
                                else:
                                    result['hostname'] = hardware_info.hostname
                                    result['output'] = '采集成功'
                                    result['status'] = 'success'

                            # 更新资产信息（无论采集成功或失败都更新状态）
                            try:
                                # 根据IP地址查找对应的Host对象
                                host = Host.objects.get(ip=ip)
                                # 更新资产信息
                                if task.update_hostname:
                                    host.hostname = hardware_info.hostname
                                if task.update_os:
                                    host.os = f"{hardware_info.os} {hardware_info.os_version}".strip()
                                if task.update_cpu:
                                    host.cpu_model = hardware_info.cpu_model
                                    host.cpu_num = str(hardware_info.cpu_num)
                                    host.cpu_cores = str(hardware_info.cpu_cores)
                                if task.update_memory:
                                    host.memory = hardware_info.memory_total
                                if task.update_disk:
                                    host.disk = hardware_info.disk_info
                                if task.update_gpu:
                                    host.gpu_model = hardware_info.gpu_model
                                if task.update_sn:
                                    host.sn = hardware_info.sn
                                if task.update_device_info:
                                    host.device_model = f"{hardware_info.vendor} {hardware_info.product}".strip()
                                    host.bm_ip = hardware_info.bm_ip
                                    host.asset_type = hardware_info.asset_type
                                # 无论是否更新其他信息，都更新设备状态和默认值
                                if result['status'] == 'success':
                                    host.status = '1'  # 使用中
                                else:
                                    host.status = '4'  # 其他
                                # 如果部门为空，设置默认值
                                if not host.department:
                                    host.department = '互联网大数据'
                                # 如果机房为空，设置默认值
                                if not host.idc:
                                    idc = Idc.objects.filter(name='智能化104机房').first()
                                    host.idc = idc
                                host.save()
                                logger.error(f"[CollectTask] 资产 {ip} 信息已更新，状态: {'使用中' if result['status'] == 'success' else '其他'}, 部门: {host.department}, 机房: {host.idc.name if host.idc else '未设置'}")
                            except Host.DoesNotExist:
                                # 资产不存在，创建新资产
                                # 查找智能化104机房
                                idc = Idc.objects.filter(name='智能化104机房').first()
                                # 确定设备状态：采集成功为使用中，失败为其他
                                device_status = '1' if result['status'] == 'success' else '4'
                                host = Host(
                                    ip=ip,
                                    hostname=hardware_info.hostname if task.update_hostname else f"host-{ip}",
                                    os=f"{hardware_info.os} {hardware_info.os_version}".strip() if task.update_os else "",
                                    cpu_model=hardware_info.cpu_model if task.update_cpu else "",
                                    cpu_num=str(hardware_info.cpu_num) if task.update_cpu else "",
                                    cpu_cores=str(hardware_info.cpu_cores) if task.update_cpu else "",
                                    memory=hardware_info.memory_total if task.update_memory else "",
                                    disk=hardware_info.disk_info if task.update_disk else "",
                                    gpu_model=hardware_info.gpu_model if task.update_gpu else "",
                                    sn=hardware_info.sn if task.update_sn else "",
                                    device_model=f"{hardware_info.vendor} {hardware_info.product}".strip() if task.update_device_info else "",
                                    bm_ip=hardware_info.bm_ip if task.update_device_info else "",
                                    asset_type=hardware_info.asset_type if task.update_device_info else "服务器",
                                    status=device_status,
                                    department='互联网大数据',  # 默认部门
                                    idc=idc  # 默认机房
                                )
                                host.save()
                                logger.error(f"[CollectTask] 资产 {ip} 不存在，已创建新资产，状态: {'使用中' if device_status == '1' else '其他'}, 部门: 互联网大数据, 机房: 智能化104机房")
                            except Exception as e:
                                logger.error(f"[CollectTask] 更新资产 {ip} 信息失败: {str(e)}")

                    # 保存历史记录
                    CollectHistory.objects.create(
                        task=task,
                        host_ip=ip,
                        hostname=result['hostname'],
                        status=result['status'],
                        error_message=result['error']
                    )

                except Exception as e:
                    result['error'] = str(e)
                    result['status'] = 'failed'
                    # 保存失败记录
                    CollectHistory.objects.create(
                        task=task,
                        host_ip=ip,
                        status='failed',
                        error_message=str(e)
                    )

                progress_data['results'].append(result)
                progress_data['completed'] += 1

                # 更新进度
                with open(progress_path, 'w') as f:
                    json.dump(progress_data, f)

            # 检查是否被停止
            try:
                with open(progress_path, 'r') as f:
                    progress_data = json.load(f)
                if progress_data.get('stopped', False):
                    # 保持停止状态
                    logger.error(f"[CollectTask] 任务已停止，保持停止状态")
                else:
                    # 完成
                    progress_data['status'] = 'completed'
                    progress_data['finished'] = True
                    with open(progress_path, 'w') as f:
                        json.dump(progress_data, f)
                    logger.error(f"[CollectTask] 任务已完成")
            except:
                pass

        # 启动线程执行采集
        thread = threading.Thread(target=execute_collection)
        thread.daemon = True
        thread.start()

        # 计算总主机数
        total_hosts = 0
        if task.target_group:
            total_hosts += task.target_group.serverList.count()
        if task.target_hosts:
            total_hosts += len(parse_ip_range(task.target_hosts))

        # 跳转到进度页面
        return render(request, 'cmdb/collect_task_progress.html', {
            'task': task,
            'total_hosts': total_hosts,
            'progress_file': progress_file
        })
    except Exception as e:
        logger.error(f"[CollectTask] 执行失败: {str(e)}")
        return JsonResponse({'success': False, 'message': f'采集任务功能暂不可用，请联系管理员。错误: {str(e)}'})

@login_required
def immediate_collect(request, task_id):
    import logging
    import time
    import threading
    from django.shortcuts import render
    logger = logging.getLogger(__name__)

    try:
        logger.error(f"[CollectTask] 立即执行采集任务 {task_id}")

        if not CollectTask or not CollectHistory:
            raise Exception('模型未导入')

        task = get_object_or_404(CollectTask, id=task_id)
        logger.error(f"[CollectTask] 找到采集任务: {task.name}, 目标主机: {task.target_hosts}")
        logger.error(f"[CollectTask] task类型: {type(task)}, task_id类型: {type(task_id)}")

        progress_file = f'collect_task_progress_{task_id}_{int(time.time())}.json'
        logger.error(f"[CollectTask] 进度文件: {progress_file}")

        def execute_collection():
            import json
            import os
            from datetime import datetime

            logger.error(f"[CollectTask] execute_collection 开始执行")

            progress_data = {
                'task_id': task_id,
                'total': 0,
                'completed': 0,
                'status': 'running',
                'stopped': False,
                'finished': False,
                'results': []
            }

            temp_dir = '/tmp'
            progress_path = os.path.join(temp_dir, progress_file)

            # 保存初始进度
            with open(progress_path, 'w') as f:
                json.dump(progress_data, f)

            logger.error(f"[CollectTask] 进度文件已创建: {progress_path}")

            # 解析目标主机
            target_ips = []
            if task.target_group:
                # 从资产组获取主机
                hosts = task.target_group.serverList.all()
                target_ips.extend([host.ip for host in hosts])
                logger.error(f"[CollectTask] 从资产组获取主机: {[host.ip for host in hosts]}")

            if task.target_hosts:
                # 解析IP范围或逗号分隔的IP列表
                parsed_ips = parse_ip_range(task.target_hosts)
                logger.error(f"[CollectTask] 解析目标主机: {task.target_hosts} -> {parsed_ips}")
                target_ips.extend(parsed_ips)

            # 去重并按IP地址排序
            target_ips = list(set(target_ips))
            # 按IP地址排序，确保从192.168.12.1开始
            target_ips.sort(key=lambda ip: tuple(map(int, ip.split('.'))))
            progress_data['total'] = len(target_ips)

            # 保存更新的进度
            with open(progress_path, 'w') as f:
                json.dump(progress_data, f)

            # 执行采集
            for ip in target_ips:
                # 检查是否停止
                try:
                    with open(progress_path, 'r') as f:
                        progress_data = json.load(f)
                    if progress_data.get('stopped', False):
                        # 标记任务为已停止
                        progress_data['stopped'] = True
                        progress_data['status'] = 'stopped'
                        progress_data['finished'] = True
                        with open(progress_path, 'w') as f:
                            json.dump(progress_data, f)
                        logger.error(f"[CollectTask] 任务已停止")
                        break
                except:
                    pass

                result = {
                    'ip': ip,
                    'hostname': '',
                    'status': 'failed',
                    'output': '',
                    'error': ''
                }

                try:
                    # 检查是否停止
                    try:
                        with open(progress_path, 'r') as f:
                            progress_data = json.load(f)
                        if progress_data.get('stopped', False):
                            # 标记任务为已停止
                            progress_data['stopped'] = True
                            progress_data['status'] = 'stopped'
                            progress_data['finished'] = True
                            with open(progress_path, 'w') as f:
                                json.dump(progress_data, f)
                            logger.error(f"[CollectTask] 任务已停止")
                            break
                    except:
                        pass

                    # 执行采集（使用真实的硬件采集工具）
                    import socket
                    
                    # 检测目标主机是否可访问
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(2)  # 2秒超时
                    result_port = sock.connect_ex((ip, 22))  # 尝试连接SSH端口
                    sock.close()
                    
                    if result_port != 0:
                        # 主机不可访问，标记为失败
                        result['error'] = f'主机 {ip} 不可访问（SSH端口22未开放）'
                        result['status'] = 'failed'
                        
                        # 更新设备状态为其他
                        try:
                            host = Host.objects.get(ip=ip)
                            host.status = '4'  # 其他
                            host.save()
                            logger.error(f"[CollectTask] 资产 {ip} 不可访问，状态更新为其他")
                        except Host.DoesNotExist:
                            # 资产不存在，不需要处理
                            pass
                        except Exception as e:
                            logger.error(f"[CollectTask] 更新资产 {ip} 状态失败: {str(e)}")
                    else:
                        # 再次检查是否停止
                        try:
                            with open(progress_path, 'r') as f:
                                progress_data = json.load(f)
                            if progress_data.get('stopped', False):
                                # 标记任务为已停止
                                progress_data['stopped'] = True
                                progress_data['status'] = 'stopped'
                                progress_data['finished'] = True
                                with open(progress_path, 'w') as f:
                                    json.dump(progress_data, f)
                                logger.error(f"[CollectTask] 任务已停止")
                                break
                        except:
                            pass

                        # 主机可访问，使用真实的硬件采集工具
                        # 查找可用的SSH配置
                        ssh_config = SSHConfig.objects.filter(is_enabled=True).first()
                        if not ssh_config:
                            result['error'] = '没有可用的SSH配置'
                            result['status'] = 'failed'
                        else:
                            # 再次检查是否停止
                            try:
                                with open(progress_path, 'r') as f:
                                    progress_data = json.load(f)
                                if progress_data.get('stopped', False):
                                    # 标记任务为已停止
                                    progress_data['stopped'] = True
                                    progress_data['status'] = 'stopped'
                                    progress_data['finished'] = True
                                    with open(progress_path, 'w') as f:
                                        json.dump(progress_data, f)
                                    logger.error(f"[CollectTask] 任务已停止")
                                    break
                            except:
                                pass

                            # 执行真实的硬件采集（使用超时控制）
                            logger.error(f"[CollectTask] 开始采集主机 {ip}")
                            
                            # 使用线程执行采集，以便可以中断
                            import threading
                            import time
                            hardware_info = None
                            采集完成 = False
                            
                            def 采集_thread():
                                nonlocal hardware_info, 采集完成
                                try:
                                    hardware_info = collect_single_server(
                                        ip, 
                                        ssh_config.port, 
                                        ssh_config.username, 
                                        ssh_config.password,
                                        collect_disk=task.update_disk
                                    )
                                finally:
                                    采集完成 = True
                            
                            # 启动采集线程
                            thread = threading.Thread(target=采集_thread)
                            thread.daemon = True
                            thread.start()
                            
                            # 等待采集完成，同时定期检查是否停止
                            start_time = time.time()
                            timeout = 30  # 30秒超时
                            while not 采集完成 and time.time() - start_time < timeout:
                                # 检查是否停止
                                try:
                                    with open(progress_path, 'r') as f:
                                        progress_data = json.load(f)
                                    if progress_data.get('stopped', False):
                                        # 标记为停止状态
                                        result['error'] = '任务已停止'
                                        result['status'] = 'stopped'
                                        # 标记任务为已停止
                                        progress_data['stopped'] = True
                                        progress_data['status'] = 'stopped'
                                        progress_data['finished'] = True
                                        with open(progress_path, 'w') as f:
                                            json.dump(progress_data, f)
                                        logger.error(f"[CollectTask] 任务已停止")
                                        break
                                except:
                                    pass
                                time.sleep(1)  # 1秒检查一次
                            
                            if not 采集完成:
                                # 采集超时
                                result['error'] = '采集超时（30秒）'
                                result['status'] = 'failed'
                            else:
                                if hardware_info.error:
                                    result['error'] = hardware_info.error
                                    result['status'] = 'failed'
                                else:
                                    result['hostname'] = hardware_info.hostname
                                    result['output'] = '采集成功'
                                    result['status'] = 'success'

                                    # 更新资产信息
                                    try:
                                        # 根据IP地址查找对应的Host对象
                                        host = Host.objects.get(ip=ip)
                                        # 更新资产信息
                                        if task.update_hostname:
                                            host.hostname = hardware_info.hostname
                                        if task.update_os:
                                            host.os = f"{hardware_info.os} {hardware_info.os_version}".strip()
                                        if task.update_cpu:
                                            host.cpu_model = hardware_info.cpu_model
                                            host.cpu_num = str(hardware_info.cpu_num)
                                            host.cpu_cores = str(hardware_info.cpu_cores)
                                        if task.update_memory:
                                            host.memory = hardware_info.memory_total
                                        if task.update_disk:
                                            host.disk = hardware_info.disk_info
                                        if task.update_gpu:
                                            host.gpu_model = hardware_info.gpu_model
                                        if task.update_sn:
                                            host.sn = hardware_info.sn
                                        if task.update_device_info:
                                            host.device_model = f"{hardware_info.vendor} {hardware_info.product}".strip()
                                            host.bm_ip = hardware_info.bm_ip
                                            host.asset_type = hardware_info.asset_type
                                        # 无论是否更新其他信息，都更新设备状态
                                        if result['status'] == 'success':
                                            host.status = '1'  # 使用中
                                        else:
                                            host.status = '4'  # 其他
                                        host.save()
                                        logger.error(f"[CollectTask] 资产 {ip} 信息已更新，状态: {'使用中' if result['status'] == 'success' else '其他'}")
                                    except Host.DoesNotExist:
                                        # 资产不存在，创建新资产
                                        # 确定设备状态：采集成功为使用中，失败为其他
                                        device_status = '1' if result['status'] == 'success' else '4'
                                        host = Host(
                                            ip=ip,
                                            hostname=hardware_info.hostname if task.update_hostname else f"host-{ip}",
                                            os=f"{hardware_info.os} {hardware_info.os_version}".strip() if task.update_os else "",
                                            cpu_model=hardware_info.cpu_model if task.update_cpu else "",
                                            cpu_num=str(hardware_info.cpu_num) if task.update_cpu else "",
                                            cpu_cores=str(hardware_info.cpu_cores) if task.update_cpu else "",
                                            memory=hardware_info.memory_total if task.update_memory else "",
                                            disk=hardware_info.disk_info if task.update_disk else "",
                                            gpu_model=hardware_info.gpu_model if task.update_gpu else "",
                                            sn=hardware_info.sn if task.update_sn else "",
                                            device_model=f"{hardware_info.vendor} {hardware_info.product}".strip() if task.update_device_info else "",
                                            bm_ip=hardware_info.bm_ip if task.update_device_info else "",
                                            asset_type=hardware_info.asset_type if task.update_device_info else "服务器",
                                            status=device_status
                                        )
                                        host.save()
                                        logger.error(f"[CollectTask] 资产 {ip} 不存在，已创建新资产，状态: {'使用中' if device_status == '1' else '其他'}")
                                    except Exception as e:
                                        logger.error(f"[CollectTask] 更新资产 {ip} 信息失败: {str(e)}")

                    # 保存历史记录
                    CollectHistory.objects.create(
                        task=task,
                        host_ip=ip,
                        hostname=result['hostname'],
                        status=result['status'],
                        error_message=result['error']
                    )

                except Exception as e:
                    result['error'] = str(e)
                    result['status'] = 'failed'
                    # 保存失败记录
                    CollectHistory.objects.create(
                        task=task,
                        host_ip=ip,
                        status='failed',
                        error_message=str(e)
                    )

                progress_data['results'].append(result)
                progress_data['completed'] += 1

                # 更新进度
                with open(progress_path, 'w') as f:
                    json.dump(progress_data, f)

            # 检查是否被停止
            try:
                with open(progress_path, 'r') as f:
                    progress_data = json.load(f)
                if progress_data.get('stopped', False):
                    # 保持停止状态
                    logger.error(f"[CollectTask] 任务已停止，保持停止状态")
                else:
                    # 完成
                    progress_data['status'] = 'completed'
                    progress_data['finished'] = True
                    with open(progress_path, 'w') as f:
                        json.dump(progress_data, f)
                    logger.error(f"[CollectTask] 任务已完成")
            except:
                pass

        # 启动线程执行采集
        thread = threading.Thread(target=execute_collection)
        thread.daemon = True
        thread.start()

        # 计算总主机数
        total_hosts = 0
        if task.target_group:
            total_hosts += task.target_group.serverList.count()
        if task.target_hosts:
            total_hosts += len(parse_ip_range(task.target_hosts))

        # 跳转到进度页面
        return render(request, 'cmdb/collect_task_progress.html', {
            'task': task,
            'total_hosts': total_hosts,
            'progress_file': progress_file
        })
    except Exception as e:
        logger.error(f"[CollectTask] 执行失败: {str(e)}")
        return JsonResponse({'success': False, 'message': f'采集任务功能暂不可用，请联系管理员。错误: {str(e)}'})

@login_required
def restart_collect_task(request, task_id):
    import logging
    import time
    import threading
    from django.shortcuts import render
    logger = logging.getLogger(__name__)

    try:
        logger.error(f"[CollectTask] 重启执行采集任务 {task_id}")

        if not CollectTask or not CollectHistory:
            raise Exception('模型未导入')

        task = get_object_or_404(CollectTask, id=task_id)
        logger.error(f"[CollectTask] 找到采集任务: {task.name}, 目标主机: {task.target_hosts}")

        progress_file = f'collect_task_progress_{task_id}_{int(time.time())}.json'
        logger.error(f"[CollectTask] 进度文件: {progress_file}")

        def execute_collection():
            import json
            import os
            from datetime import datetime

            logger.error(f"[CollectTask] execute_collection 开始执行")

            progress_data = {
                'task_id': task_id,
                'total': 0,
                'completed': 0,
                'status': 'running',
                'stopped': False,
                'finished': False,
                'results': []
            }

            temp_dir = '/tmp'
            progress_path = os.path.join(temp_dir, progress_file)

            # 保存初始进度
            with open(progress_path, 'w') as f:
                json.dump(progress_data, f)

            logger.error(f"[CollectTask] 进度文件已创建: {progress_path}")

            # 解析目标主机
            target_ips = []
            if task.target_group:
                # 从资产组获取主机
                hosts = task.target_group.serverList.all()
                target_ips.extend([host.ip for host in hosts])
                logger.error(f"[CollectTask] 从资产组获取主机: {[host.ip for host in hosts]}")

            if task.target_hosts:
                # 解析IP范围或逗号分隔的IP列表
                parsed_ips = parse_ip_range(task.target_hosts)
                logger.error(f"[CollectTask] 解析目标主机: {task.target_hosts} -> {parsed_ips}")
                target_ips.extend(parsed_ips)

            # 去重并按IP地址排序
            target_ips = list(set(target_ips))
            # 按IP地址排序，确保从192.168.12.1开始
            target_ips.sort(key=lambda ip: tuple(map(int, ip.split('.'))))
            progress_data['total'] = len(target_ips)

            # 保存更新的进度
            with open(progress_path, 'w') as f:
                json.dump(progress_data, f)

            # 执行采集
            for ip in target_ips:
                # 检查是否停止
                try:
                    with open(progress_path, 'r') as f:
                        progress_data = json.load(f)
                    if progress_data.get('stopped', False):
                        # 标记任务为已停止
                        progress_data['stopped'] = True
                        progress_data['status'] = 'stopped'
                        progress_data['finished'] = True
                        with open(progress_path, 'w') as f:
                            json.dump(progress_data, f)
                        logger.error(f"[CollectTask] 任务已停止")
                        break
                except:
                    pass

                result = {
                    'ip': ip,
                    'hostname': '',
                    'status': 'failed',
                    'output': '',
                    'error': ''
                }

                try:
                    # 检查是否停止
                    try:
                        with open(progress_path, 'r') as f:
                            progress_data = json.load(f)
                        if progress_data.get('stopped', False):
                            # 标记任务为已停止
                            progress_data['stopped'] = True
                            progress_data['status'] = 'stopped'
                            progress_data['finished'] = True
                            with open(progress_path, 'w') as f:
                                json.dump(progress_data, f)
                            logger.error(f"[CollectTask] 任务已停止")
                            break
                    except:
                        pass

                    # 执行采集（使用真实的硬件采集工具）
                    import socket
                    
                    # 检测目标主机是否可访问
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(2)  # 2秒超时
                    result_port = sock.connect_ex((ip, 22))  # 尝试连接SSH端口
                    sock.close()
                    
                    if result_port != 0:
                        # 主机不可访问，标记为失败
                        result['error'] = f'主机 {ip} 不可访问（SSH端口22未开放）'
                        result['status'] = 'failed'
                        
                        # 更新设备状态为其他
                        try:
                            host = Host.objects.get(ip=ip)
                            host.status = '4'  # 其他
                            host.save()
                            logger.error(f"[CollectTask] 资产 {ip} 不可访问，状态更新为其他")
                        except Host.DoesNotExist:
                            # 资产不存在，不需要处理
                            pass
                        except Exception as e:
                            logger.error(f"[CollectTask] 更新资产 {ip} 状态失败: {str(e)}")
                    else:
                        # 再次检查是否停止
                        try:
                            with open(progress_path, 'r') as f:
                                progress_data = json.load(f)
                            if progress_data.get('stopped', False):
                                # 标记任务为已停止
                                progress_data['stopped'] = True
                                progress_data['status'] = 'stopped'
                                progress_data['finished'] = True
                                with open(progress_path, 'w') as f:
                                    json.dump(progress_data, f)
                                logger.error(f"[CollectTask] 任务已停止")
                                break
                        except:
                            pass

                        # 主机可访问，使用真实的硬件采集工具
                        # 查找可用的SSH配置
                        ssh_config = SSHConfig.objects.filter(is_enabled=True).first()
                        if not ssh_config:
                            result['error'] = '没有可用的SSH配置'
                            result['status'] = 'failed'
                        else:
                            # 再次检查是否停止
                            try:
                                with open(progress_path, 'r') as f:
                                    progress_data = json.load(f)
                                if progress_data.get('stopped', False):
                                    # 标记任务为已停止
                                    progress_data['stopped'] = True
                                    progress_data['status'] = 'stopped'
                                    progress_data['finished'] = True
                                    with open(progress_path, 'w') as f:
                                        json.dump(progress_data, f)
                                    logger.error(f"[CollectTask] 任务已停止")
                                    break
                            except:
                                pass

                            # 执行真实的硬件采集（使用超时控制）
                            logger.error(f"[CollectTask] 开始采集主机 {ip}")
                            
                            # 使用线程执行采集，以便可以中断
                            import threading
                            import time
                            hardware_info = None
                            采集完成 = False
                            
                            def 采集_thread():
                                nonlocal hardware_info, 采集完成
                                try:
                                    hardware_info = collect_single_server(
                                        ip, 
                                        ssh_config.port, 
                                        ssh_config.username, 
                                        ssh_config.password,
                                        collect_disk=task.update_disk
                                    )
                                finally:
                                    采集完成 = True
                            
                            # 启动采集线程
                            thread = threading.Thread(target=采集_thread)
                            thread.daemon = True
                            thread.start()
                            
                            # 等待采集完成，同时定期检查是否停止
                            start_time = time.time()
                            timeout = 30  # 30秒超时
                            while not 采集完成 and time.time() - start_time < timeout:
                                # 检查是否停止
                                try:
                                    with open(progress_path, 'r') as f:
                                        progress_data = json.load(f)
                                    if progress_data.get('stopped', False):
                                        # 标记为停止状态
                                        result['error'] = '任务已停止'
                                        result['status'] = 'stopped'
                                        # 标记任务为已停止
                                        progress_data['stopped'] = True
                                        progress_data['status'] = 'stopped'
                                        progress_data['finished'] = True
                                        with open(progress_path, 'w') as f:
                                            json.dump(progress_data, f)
                                        logger.error(f"[CollectTask] 任务已停止")
                                        break
                                except:
                                    pass
                                time.sleep(1)  # 1秒检查一次
                            
                            if not 采集完成:
                                # 采集超时
                                result['error'] = '采集超时（30秒）'
                                result['status'] = 'failed'
                            else:
                                if hardware_info.error:
                                    result['error'] = hardware_info.error
                                    result['status'] = 'failed'
                                else:
                                    result['hostname'] = hardware_info.hostname
                                    result['output'] = '采集成功'
                                    result['status'] = 'success'

                                    # 更新资产信息
                                    try:
                                        # 根据IP地址查找对应的Host对象
                                        host = Host.objects.get(ip=ip)
                                        # 更新资产信息
                                        if task.update_hostname:
                                            host.hostname = hardware_info.hostname
                                        if task.update_os:
                                            host.os = f"{hardware_info.os} {hardware_info.os_version}".strip()
                                        if task.update_cpu:
                                            host.cpu_model = hardware_info.cpu_model
                                            host.cpu_num = str(hardware_info.cpu_num)
                                            host.cpu_cores = str(hardware_info.cpu_cores)
                                        if task.update_memory:
                                            host.memory = hardware_info.memory_total
                                        if task.update_disk:
                                            host.disk = hardware_info.disk_info
                                        if task.update_gpu:
                                            host.gpu_model = hardware_info.gpu_model
                                        if task.update_sn:
                                            host.sn = hardware_info.sn
                                        if task.update_device_info:
                                            host.device_model = f"{hardware_info.vendor} {hardware_info.product}".strip()
                                            host.bm_ip = hardware_info.bm_ip
                                            host.asset_type = hardware_info.asset_type
                                        host.save()
                                        logger.error(f"[CollectTask] 资产 {ip} 信息已更新")
                                    except Host.DoesNotExist:
                                        # 资产不存在，创建新资产
                                        host = Host(
                                            ip=ip,
                                            hostname=hardware_info.hostname if task.update_hostname else f"host-{ip}",
                                            os=f"{hardware_info.os} {hardware_info.os_version}".strip() if task.update_os else "",
                                            cpu_model=hardware_info.cpu_model if task.update_cpu else "",
                                            cpu_num=str(hardware_info.cpu_num) if task.update_cpu else "",
                                            cpu_cores=str(hardware_info.cpu_cores) if task.update_cpu else "",
                                            memory=hardware_info.memory_total if task.update_memory else "",
                                            disk=hardware_info.disk_info if task.update_disk else "",
                                            gpu_model=hardware_info.gpu_model if task.update_gpu else "",
                                            sn=hardware_info.sn if task.update_sn else "",
                                            device_model=f"{hardware_info.vendor} {hardware_info.product}".strip() if task.update_device_info else "",
                                            bm_ip=hardware_info.bm_ip if task.update_device_info else "",
                                            asset_type=hardware_info.asset_type if task.update_device_info else "服务器"
                                        )
                                        host.save()
                                        logger.error(f"[CollectTask] 资产 {ip} 不存在，已创建新资产")
                                    except Exception as e:
                                        logger.error(f"[CollectTask] 更新资产 {ip} 信息失败: {str(e)}")

                    # 保存历史记录
                    CollectHistory.objects.create(
                        task=task,
                        host_ip=ip,
                        hostname=result['hostname'],
                        status=result['status'],
                        error_message=result['error']
                    )

                except Exception as e:
                    result['error'] = str(e)
                    result['status'] = 'failed'
                    # 保存失败记录
                    CollectHistory.objects.create(
                        task=task,
                        host_ip=ip,
                        status='failed',
                        error_message=str(e)
                    )

                progress_data['results'].append(result)
                progress_data['completed'] += 1

                # 更新进度
                with open(progress_path, 'w') as f:
                    json.dump(progress_data, f)

            # 检查是否被停止
            try:
                with open(progress_path, 'r') as f:
                    progress_data = json.load(f)
                if progress_data.get('stopped', False):
                    # 保持停止状态
                    logger.error(f"[CollectTask] 任务已停止，保持停止状态")
                else:
                    # 完成
                    progress_data['status'] = 'completed'
                    progress_data['finished'] = True
                    with open(progress_path, 'w') as f:
                        json.dump(progress_data, f)
                    logger.error(f"[CollectTask] 任务已完成")
            except:
                pass

        # 启动线程执行采集
        thread = threading.Thread(target=execute_collection)
        thread.daemon = True
        thread.start()

        # 计算总主机数
        total_hosts = 0
        if task.target_group:
            total_hosts += task.target_group.serverList.count()
        if task.target_hosts:
            total_hosts += len(parse_ip_range(task.target_hosts))

        # 跳转到进度页面
        return render(request, 'cmdb/collect_task_progress.html', {
            'task': task,
            'total_hosts': total_hosts,
            'progress_file': progress_file
        })
    except Exception as e:
        logger.error(f"[CollectTask] 执行失败: {str(e)}")
        return JsonResponse({'success': False, 'message': f'采集任务功能暂不可用，请联系管理员。错误: {str(e)}'})

# 停止采集任务

@login_required
def stop_collect_task(request, task_id):
    try:
        # 查找正在执行的采集任务进度文件
        import glob
        import os
        import json
        import logging
        logger = logging.getLogger(__name__)
        
        temp_dir = "/tmp"
        progress_files = glob.glob(os.path.join(temp_dir, f"collect_task_progress_{task_id}_*.json"))
        
        if not progress_files:
            return JsonResponse({"success": False, "message": "未找到采集任务进度文件"}, json_dumps_params={'ensure_ascii': False})
        
        # 按修改时间排序，获取最新的进度文件
        progress_files.sort(key=os.path.getmtime, reverse=True)
        progress_path = progress_files[0]
        
        # 读取进度文件
        with open(progress_path, "r") as f:
            progress_data = json.load(f)
        
        # 检查任务是否已完成
        if progress_data.get("finished", False):
            return JsonResponse({"success": False, "message": "任务已完成，无需停止"}, json_dumps_params={'ensure_ascii': False})
        
        # 设置停止标志
        progress_data["stopped"] = True
        progress_data["status"] = "stopped"
        
        # 保存更新后的进度文件
        with open(progress_path, "w") as f:
            json.dump(progress_data, f)
        
        logger.error(f"[CollectTask] 任务 {task_id} 已标记为停止")
        return JsonResponse({"success": True, "message": "任务已成功停止"}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"[CollectTask] 停止任务失败: {str(e)}")
        return JsonResponse({"success": False, "message": f"停止任务失败: {str(e)}"}, json_dumps_params={'ensure_ascii': False})

# 批量命令相关视图
@login_required
def batch_command_list(request):
    try:
        commands = BatchCommand.objects.all()
        return render(request, 'cmdb/batch_command_list.html', {'commands': commands})
    except Exception as e:
        messages.error(request, f'获取批量命令列表失败: {str(e)}')
        return render(request, 'cmdb/batch_command_list.html', {'commands': []})

@login_required
def batch_command_add(request):
    try:
        return render(request, 'cmdb/batch_command_form.html', {'action': 'add'})
    except Exception as e:
        messages.error(request, f'添加批量命令失败: {str(e)}')
        return redirect('batch_command_list')

@login_required
def batch_command_edit(request, command_id):
    try:
        command = get_object_or_404(BatchCommand, id=command_id)
        return render(request, 'cmdb/batch_command_form.html', {'command': command, 'action': 'edit'})
    except Exception as e:
        messages.error(request, f'编辑批量命令失败: {str(e)}')
        return redirect('batch_command_list')

@login_required
def batch_command_delete(request, command_id):
    try:
        command = get_object_or_404(BatchCommand, id=command_id)
        command.delete()
        messages.success(request, f'批量命令 {command.name} 删除成功')
        return redirect('batch_command_list')
    except Exception as e:
        messages.error(request, f'删除批量命令失败: {str(e)}')
        return redirect('batch_command_list')

@login_required
def run_batch_command(request, command_id):
    try:
        command = get_object_or_404(BatchCommand, id=command_id)
        messages.success(request, f'批量命令 {command.name} 开始执行')
        return redirect('batch_command_list')
    except Exception as e:
        messages.error(request, f'执行批量命令失败: {str(e)}')
        return redirect('batch_command_list')

@login_required
def stop_batch_command(request, command_id):
    try:
        command = get_object_or_404(BatchCommand, id=command_id)
        messages.success(request, f'批量命令 {command.name} 已停止')
        return redirect('batch_command_list')
    except Exception as e:
        messages.error(request, f'停止批量命令失败: {str(e)}')
        return redirect('batch_command_list')

@login_required
def batch_command_progress(request, progress_file):
    import os
    import json
    
    try:
        temp_dir = '/tmp'
        progress_path = os.path.join(temp_dir, progress_file)
        
        if os.path.exists(progress_path):
            with open(progress_path, 'r') as f:
                progress_data = json.load(f)
            return JsonResponse(progress_data)
        else:
            return JsonResponse({'error': 'Progress file not found'})
    except Exception as e:
        return JsonResponse({'error': str(e)})


# 静态资产相关视图
@login_required
def static_asset_list(request):
    try:
        # 获取过滤参数
        keyword = request.GET.get('keyword', '')
        cabinet_filter = request.GET.get('cabinet', '')
        department_filter = request.GET.get('department', '')
        server_type_filter = request.GET.get('server_type', '')
        status_filter = request.GET.get('status', '')
        
        # 构建查询
        queryset = StaticAsset.objects.all()
        
        # 关键词搜索
        if keyword:
            queryset = queryset.filter(
                Q(serial_number__icontains=keyword) |
                Q(asset_no__icontains=keyword) |
                Q(cabinet__icontains=keyword) |
                Q(department__icontains=keyword) |
                Q(server_type__icontains=keyword) |
                Q(ip__icontains=keyword) |
                Q(contact_person__icontains=keyword) |
                Q(device_model__icontains=keyword) |
                Q(server_model__icontains=keyword) |
                Q(status__icontains=keyword)
            )
        
        # 机柜过滤
        if cabinet_filter:
            queryset = queryset.filter(cabinet__icontains=cabinet_filter)
        
        # 部门过滤
        if department_filter:
            queryset = queryset.filter(department__icontains=department_filter)
        
        # 服务器类型过滤
        if server_type_filter:
            queryset = queryset.filter(server_type__icontains=server_type_filter)
        
        # 状态过滤
        if status_filter:
            queryset = queryset.filter(status__icontains=status_filter)
        
        # 按机柜升序、起始U位降序排序（从42U到01U），无机柜的排在最后
        assets = queryset.annotate(
            has_cabinet=Case(
                When(cabinet='', then=Value(1)),
                When(cabinet__isnull=True, then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            ),
            start_u_int=Cast('start_u', IntegerField())
        ).order_by('has_cabinet', 'cabinet', '-start_u_int')
        
        # 获取过滤选项（从现有数据中提取，处理空值情况）
        try:
            cabinet_list = list(StaticAsset.objects.values_list('cabinet', flat=True).distinct().order_by('cabinet'))
            cabinet_list = [c for c in cabinet_list if c]  # 过滤掉空值
        except:
            cabinet_list = []
            
        try:
            department_list = list(StaticAsset.objects.values_list('department', flat=True).distinct().order_by('department'))
            department_list = [d for d in department_list if d]  # 过滤掉空值
        except:
            department_list = []
            
        try:
            server_type_list = list(StaticAsset.objects.values_list('server_type', flat=True).distinct().order_by('server_type'))
            server_type_list = [s for s in server_type_list if s]  # 过滤掉空值
        except:
            server_type_list = []
            
        try:
            status_list = list(StaticAsset.objects.values_list('status', flat=True).distinct().order_by('status'))
            status_list = [s for s in status_list if s]  # 过滤掉空值
        except:
            status_list = []
        
        # 计算统计数据
        total_count = StaticAsset.objects.count()
        filtered_count = assets.count()
        has_filter = keyword or cabinet_filter or department_filter or server_type_filter or status_filter
        
        response = render(request, 'cmdb/static_asset_list.html', {
            'assets': assets,
            'keyword': keyword,
            'cabinet_filter': cabinet_filter,
            'department_filter': department_filter,
            'server_type_filter': server_type_filter,
            'status_filter': status_filter,
            'cabinet_list': cabinet_list,
            'total_count': total_count,
            'filtered_count': filtered_count,
            'has_filter': has_filter,
            'department_list': department_list,
            'server_type_list': server_type_list,
            'status_list': status_list
        })
        
        # 添加缓存控制头，防止浏览器缓存页面
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        
        return response
    except Exception as e:
        messages.error(request, f'获取静态资产列表失败: {str(e)}')
        return render(request, 'cmdb/static_asset_list.html', {'assets': []})

@login_required
def static_asset_add(request):
    try:
        if request.method == 'POST':
            # 处理表单提交
            asset = StaticAsset()
            asset.serial_number = request.POST.get('serial_number')
            asset.asset_no = request.POST.get('asset_no')
            asset.cabinet = request.POST.get('cabinet')
            asset.start_u = request.POST.get('start_u')
            asset.end_u = request.POST.get('end_u')
            asset.department = request.POST.get('department')
            asset.server_type = request.POST.get('server_type')
            asset.ip = request.POST.get('ip')
            asset.contact_person = request.POST.get('contact_person')
            asset.device_model = request.POST.get('device_model')
            asset.server_model = request.POST.get('server_model')
            asset.power_rating = request.POST.get('power_rating')
            asset.memo = request.POST.get('memo')
            asset.status = request.POST.get('status')
            asset.save()

            log_operation(request.user, 'add', f'静态资产: {asset.asset_no}', f'添加静态资产: {asset.asset_no}', request.META.get('REMOTE_ADDR'))

            messages.success(request, f'静态资产 {asset.asset_no} 添加成功')
            return redirect('static_asset_list')
        else:
            # 渲染表单
            return render(request, 'cmdb/static_asset_form.html', {'action': 'add'})
    except Exception as e:
        messages.error(request, f'添加静态资产失败: {str(e)}')
        return redirect('static_asset_list')

@login_required
def static_asset_edit(request, asset_id):
    try:
        asset = get_object_or_404(StaticAsset, id=asset_id)
        
        if request.method == 'POST':
            # 处理表单提交
            asset.serial_number = request.POST.get('serial_number')
            asset.asset_no = request.POST.get('asset_no')
            asset.cabinet = request.POST.get('cabinet')
            asset.start_u = request.POST.get('start_u')
            asset.end_u = request.POST.get('end_u')
            asset.department = request.POST.get('department')
            asset.server_type = request.POST.get('server_type')
            asset.ip = request.POST.get('ip')
            asset.contact_person = request.POST.get('contact_person')
            asset.device_model = request.POST.get('device_model')
            asset.server_model = request.POST.get('server_model')
            asset.power_rating = request.POST.get('power_rating')
            asset.memo = request.POST.get('memo')
            asset.status = request.POST.get('status')
            asset.save()

            log_operation(request.user, 'edit', f'静态资产: {asset.asset_no}', f'更新静态资产: {asset.asset_no}', request.META.get('REMOTE_ADDR'))

            messages.success(request, f'静态资产 {asset.asset_no} 更新成功')
            return redirect('static_asset_list')
        else:
            # 渲染表单
            return render(request, 'cmdb/static_asset_form.html', {'asset': asset, 'action': 'edit'})
    except Exception as e:
        messages.error(request, f'编辑静态资产失败: {str(e)}')
        return redirect('static_asset_list')

@login_required
def static_asset_delete(request, asset_id):
    try:
        asset = get_object_or_404(StaticAsset, id=asset_id)
        asset_no = asset.asset_no
        asset.delete()

        log_operation(request.user, 'delete', f'静态资产: {asset_no}', f'删除静态资产: {asset_no}', request.META.get('REMOTE_ADDR'))

        messages.success(request, f'静态资产 {asset_no} 删除成功')
        return redirect('static_asset_list')
    except Exception as e:
        messages.error(request, f'删除静态资产失败: {str(e)}')
        return redirect('static_asset_list')

@login_required
def static_asset_batch_delete(request):
    try:
        if request.method == 'POST':
            asset_ids = request.POST.getlist('asset_ids[]')
            if asset_ids:
                deleted_count = 0
                for asset_id in asset_ids:
                    try:
                        asset = get_object_or_404(StaticAsset, id=asset_id)
                        asset.delete()
                        deleted_count += 1
                    except:
                        continue

                log_operation(request.user, 'delete', f'静态资产(批量)', f'批量删除 {deleted_count} 条静态资产', request.META.get('REMOTE_ADDR'))

                messages.success(request, f'成功删除 {deleted_count} 条静态资产')
            return redirect('static_asset_list')
        return redirect('static_asset_list')
    except Exception as e:
        messages.error(request, f'批量删除静态资产失败: {str(e)}')
        return redirect('static_asset_list')

@login_required
def static_asset_import(request):
    try:
        if request.method == 'POST':
            # 处理文件上传
            if 'file' not in request.FILES:
                messages.error(request, '请选择要导入的Excel文件')
                return redirect('static_asset_import')
            
            file = request.FILES['file']
            
            # 检查文件类型
            if not file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, '请上传Excel文件（.xlsx 或 .xls格式）')
                return redirect('static_asset_import')
            
            # 读取Excel文件
            import openpyxl
            from io import BytesIO
            
            # 加载文件
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            # 验证表头 - 更灵活的匹配方式
            expected_headers = ['序号', '服务器资产编号', '机柜', '开始U数', '结束U数', '使用部门/团队（负责人）', '服务器类型', 'IP地址', '联系人/责任人/使用人', '服务器品牌型号', '服务器机型（CPU/GPU卡型数量）', '服务器额定功率', '备注', '状态']
            
            # 清理表头，移除空格和特殊字符
            cleaned_headers = []
            for header in headers:
                # 移除空格和特殊字符
                cleaned = ''.join(e for e in header if e.isalnum() or e in ['(', ')', '/', '（', '）', ':', '：'])
                cleaned_headers.append(cleaned)
            
            # 清理预期表头
            cleaned_expected = []
            for header in expected_headers:
                cleaned = ''.join(e for e in header if e.isalnum() or e in ['(', ')', '/', '（', '）', ':', '：'])
                cleaned_expected.append(cleaned)
            
            # 检查是否包含所有必要的列
            missing_headers = []
            for expected in cleaned_expected:
                found = False
                for actual in cleaned_headers:
                    if expected in actual or actual in expected:
                        found = True
                        break
                if not found:
                    missing_headers.append(expected)
            
            if missing_headers:
                messages.error(request, f'Excel文件缺少必要的列：{missing_headers}')
                return redirect('static_asset_import')
            
            # 解析数据
            imported_count = 0
            skipped_count = 0
            error_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 跳过空行
                if not any(row):
                    skipped_count += 1
                    continue
                
                try:
                    # 构建资产数据
                    def get_column_value(expected_header):
                        # 清理预期表头
                        cleaned_expected = ''.join(e for e in expected_header if e.isalnum() or e in ['(', ')', '/', '（', '）', ':', '：'])
                        # 查找匹配的列
                        for i, header in enumerate(headers):
                            cleaned_actual = ''.join(e for e in header if e.isalnum() or e in ['(', ')', '/', '（', '）', ':', '：'])
                            if cleaned_expected in cleaned_actual or cleaned_actual in cleaned_expected:
                                if i < len(row) and row[i] is not None:
                                    return str(row[i])
                                return ''
                        return ''
                    
                    asset_data = {
                        'serial_number': get_column_value('序号'),
                        'asset_no': get_column_value('服务器资产编号'),
                        'cabinet': get_column_value('机柜'),
                        'start_u': get_column_value('开始U数'),
                        'end_u': get_column_value('结束U数'),
                        'department': get_column_value('使用部门/团队（负责人）'),
                        'server_type': get_column_value('服务器类型'),
                        'ip': get_column_value('IP地址') or None,
                        'contact_person': get_column_value('联系人/责任人/使用人'),
                        'device_model': get_column_value('服务器品牌型号'),
                        'server_model': get_column_value('服务器机型（CPU/GPU卡型数量）'),
                        'power_rating': get_column_value('服务器额定功率'),
                        'memo': get_column_value('备注'),
                        'status': get_column_value('状态')
                    }
                    
                    # 验证必要字段
                    if not asset_data['asset_no']:
                        skipped_count += 1
                        continue
                    
                    # 查找现有资产
                    existing_asset = StaticAsset.objects.filter(asset_no=asset_data['asset_no']).first()
                    
                    if existing_asset:
                        # 更新现有资产
                        for key, value in asset_data.items():
                            setattr(existing_asset, key, value)
                        existing_asset.save()
                    else:
                        # 创建新资产
                        StaticAsset.objects.create(**asset_data)
                    
                    imported_count += 1
                    
                except Exception as e:
                    error_count += 1
                    continue
            
            # 显示导入结果
            log_operation(request.user, 'import', f'静态资产', f'导入静态资产Excel: 成功{imported_count}条, 跳过{skipped_count}条, 错误{error_count}条', request.META.get('REMOTE_ADDR'))

            messages.success(request, f'导入完成！成功导入 {imported_count} 条数据，跳过 {skipped_count} 条，错误 {error_count} 条')
            return redirect('static_asset_list')
        else:
            # 渲染导入页面
            return render(request, 'cmdb/static_asset_import.html')
    except Exception as e:
        messages.error(request, f'导入静态资产失败: {str(e)}')
        return redirect('static_asset_list')

@login_required
def export_static_assets_excel(request):
    try:
        import io
        from openpyxl import Workbook
        from datetime import datetime
        
        # 获取过滤参数（支持从列表页传递过滤条件）
        keyword = request.GET.get('keyword', '')
        cabinet_filter = request.GET.get('cabinet', '')
        department_filter = request.GET.get('department', '')
        server_type_filter = request.GET.get('server_type', '')
        status_filter = request.GET.get('status', '')
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '静态资产列表'
        
        # 添加表头
        headers = ['序号', '服务器资产编号', '机柜', '开始U数', '结束U数', '使用部门/团队（负责人）', '服务器类型', 'IP地址', '联系人/责任人/使用人', '服务器品牌型号', '服务器机型（CPU/GPU卡型数量）', '服务器额定功率', '备注', '状态']
        ws.append(headers)
        
        # 获取静态资产数据
        assets = StaticAsset.objects.all()
        
        # 应用过滤条件
        from django.db.models import Q
        if keyword:
            assets = assets.filter(
                Q(serial_number__icontains=keyword) |
                Q(asset_no__icontains=keyword) |
                Q(cabinet__icontains=keyword) |
                Q(department__icontains=keyword) |
                Q(server_type__icontains=keyword) |
                Q(ip__icontains=keyword) |
                Q(contact_person__icontains=keyword) |
                Q(device_model__icontains=keyword) |
                Q(server_model__icontains=keyword) |
                Q(status__icontains=keyword)
            )
        
        if cabinet_filter:
            assets = assets.filter(cabinet__icontains=cabinet_filter)
        
        if department_filter:
            assets = assets.filter(department__icontains=department_filter)
        
        if server_type_filter:
            assets = assets.filter(server_type__icontains=server_type_filter)
        
        if status_filter:
            assets = assets.filter(status__icontains=status_filter)
        for asset in assets:
            row = [
                asset.serial_number or '',
                asset.asset_no or '',
                asset.cabinet or '',
                asset.start_u or '',
                asset.end_u or '',
                asset.department or '',
                asset.server_type or '',
                asset.ip or '',
                asset.contact_person or '',
                asset.device_model or '',
                asset.server_model or '',
                asset.power_rating or '',
                asset.memo or '',
                asset.status or ''
            ]
            ws.append(row)
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f'static_assets_export_{timestamp}.xlsx'
        
        # 设置响应头
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        log_operation(request.user, 'export', '静态资产', '导出静态资产Excel', request.META.get('REMOTE_ADDR'))

        return response
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'导出失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})

@login_required
def export_static_assets_csv(request):
    try:
        import csv
        import io
        from datetime import datetime
        
        # 获取过滤参数（支持从列表页传递过滤条件）
        keyword = request.GET.get('keyword', '')
        cabinet_filter = request.GET.get('cabinet', '')
        department_filter = request.GET.get('department', '')
        server_type_filter = request.GET.get('server_type', '')
        status_filter = request.GET.get('status', '')
        
        # 创建内存文件
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 添加表头
        headers = ['序号', '服务器资产编号', '机柜', '开始U数', '结束U数', '使用部门/团队（负责人）', '服务器类型', 'IP地址', '联系人/责任人/使用人', '服务器品牌型号', '服务器机型（CPU/GPU卡型数量）', '服务器额定功率', '备注', '状态']
        writer.writerow(headers)
        
        # 获取静态资产数据
        assets = StaticAsset.objects.all()
        
        # 应用过滤条件
        from django.db.models import Q
        if keyword:
            assets = assets.filter(
                Q(serial_number__icontains=keyword) |
                Q(asset_no__icontains=keyword) |
                Q(cabinet__icontains=keyword) |
                Q(department__icontains=keyword) |
                Q(server_type__icontains=keyword) |
                Q(ip__icontains=keyword) |
                Q(contact_person__icontains=keyword) |
                Q(device_model__icontains=keyword) |
                Q(server_model__icontains=keyword) |
                Q(status__icontains=keyword)
            )
        
        if cabinet_filter:
            assets = assets.filter(cabinet__icontains=cabinet_filter)
        
        if department_filter:
            assets = assets.filter(department__icontains=department_filter)
        
        if server_type_filter:
            assets = assets.filter(server_type__icontains=server_type_filter)
        
        if status_filter:
            assets = assets.filter(status__icontains=status_filter)
        for asset in assets:
            row = [
                asset.serial_number or '',
                asset.asset_no or '',
                asset.cabinet or '',
                asset.start_u or '',
                asset.end_u or '',
                asset.department or '',
                asset.server_type or '',
                asset.ip or '',
                asset.contact_person or '',
                asset.device_model or '',
                asset.server_model or '',
                asset.power_rating or '',
                asset.memo or '',
                asset.status or ''
            ]
            writer.writerow(row)
        
        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f'static_assets_export_{timestamp}.csv'
        
        # 设置响应头
        from django.http import HttpResponse
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'导出失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})

@login_required
def cabinet_layout(request):
    try:
        # 获取所有静态资产，按机柜升序排序，每个机柜内按起始U位降序排序（从42U到01U）
        assets = StaticAsset.objects.exclude(cabinet='').exclude(start_u='').exclude(end_u='').order_by('cabinet', '-start_u')
        
        # 按机柜分组
        cabinet_data = {}
        for asset in assets:
            if asset.cabinet not in cabinet_data:
                cabinet_data[asset.cabinet] = []
            cabinet_data[asset.cabinet].append(asset)
        
        # 生成机柜布局数据
        cabinets = []
        for cabinet_name in sorted(cabinet_data.keys()):
            assets_in_cabinet = cabinet_data[cabinet_name]
            
            # 创建42个U位的列表
            u_positions = {}
            for u in range(1, 43):
                u_positions[u] = None
            
            # 处理设备列表
            devices = []
            for asset in assets_in_cabinet:
                try:
                    start_u = int(asset.start_u) if asset.start_u else 1
                    end_u = int(asset.end_u) if asset.end_u else start_u
                    
                    # 确保U位在有效范围内
                    start_u = max(1, min(42, start_u))
                    end_u = max(1, min(42, end_u))
                    
                    height = end_u - start_u + 1
                    
                    # 确定设备类型
                    device_type = 'cpu'
                    server_type_lower = (asset.server_type or '').lower()
                    if 'gpu' in server_type_lower or 'GPU' in (asset.server_type or ''):
                        device_type = 'gpu'
                    elif '网络' in (asset.server_type or '') or '交换机' in (asset.server_type or '') or 'network' in server_type_lower:
                        device_type = 'network'
                    
                    # 计算grid-row值（从start_u开始，网格倒序：grid-row 1=U42，grid-row 42=U1）
                    grid_row_start = 43 - start_u
                    
                    # 添加设备信息
                    devices.append({
                        'asset_no': asset.asset_no,
                        'ip': asset.ip,
                        'start_u': start_u,
                        'end_u': end_u,
                        'height': height,
                        'type': device_type,
                        'power': asset.power_rating,
                        'device_model': asset.device_model,
                        'grid_row_start': grid_row_start,
                        'u_list': list(range(start_u, end_u + 1)),
                        'status': asset.status
                    })
                    
                    # 填充U位
                    for u in range(start_u, end_u + 1):
                        if u_positions[u] is None:
                            u_positions[u] = asset.ip if asset.ip else asset.asset_no
                except (ValueError, TypeError):
                    continue
            
            # 转换为列表格式（42U到1U，倒序）
            u_list = []
            for u in range(42, 0, -1):
                u_list.append({
                    'u': u,
                    'content': u_positions[u]
                })
            
            cabinets.append({
                'name': cabinet_name,
                'u_list': u_list,
                'assets': devices
            })
        
        return render(request, 'cmdb/cabinet_layout.html', {
            'cabinets': cabinets
        })
    except Exception as e:
        messages.error(request, f'获取机柜布局失败: {str(e)}')
        return render(request, 'cmdb/cabinet_layout.html', {'cabinets': []})


@login_required
def api_assets(request):
    assets = StaticAsset.objects.exclude(cabinet='').exclude(start_u='').exclude(end_u='')
    data = []
    for asset in assets:
        try:
            data.append({
                'cabinet': asset.cabinet,
                'asset_no': asset.asset_no,
                'model': asset.device_model,
                'ip': str(asset.ip) if asset.ip else '',
                'power': asset.power_rating,
                'start_u': int(asset.start_u),
                'end_u': int(asset.end_u),
                'status': asset.status
            })
        except:
            continue
    return JsonResponse(data, safe=False)


# 数据备份相关视图
from django.conf import settings
import subprocess
import gzip
import shutil


def get_backup_config():
    """获取备份配置，从配置文件读取，默认值基于Django settings"""
    # 配置文件保存在项目目录下，不依赖备份目录
    config_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'backup_config.json')
    
    default_config = {
        'db_host': settings.DATABASES['default']['HOST'],
        'db_port': settings.DATABASES['default']['PORT'],
        'db_user': settings.DATABASES['default']['USER'],
        'db_password': settings.DATABASES['default']['PASSWORD'],
        'db_name': settings.DATABASES['default']['NAME'],
        'backup_dir': '/data01/db_backup',
        'keep_count': 7,
        'auto_backup_enabled': False,
        'auto_backup_time': '02:00',
        'auto_backup_cron': '0 2 * * *',
        'backup_media_enabled': True
    }
    
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
                # 合并默认值
                for key, value in default_config.items():
                    if key not in config:
                        config[key] = value
                return config
        except Exception as e:
            print(f"读取备份配置失败: {str(e)}")
    
    return default_config


def save_backup_config(config):
    """保存备份配置到文件"""
    # 配置文件保存在项目目录下，不依赖备份目录
    config_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'backup_config.json')
    os.makedirs(os.path.dirname(config_file), exist_ok=True)
    
    try:
        with open(config_file, 'w') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True, None
    except Exception as e:
        return False, str(e)


@login_required
def get_backup_config_api(request):
    """API: 获取备份配置"""
    config = get_backup_config()
    # 隐藏密码字段
    config_with_masked_pwd = config.copy()
    if config_with_masked_pwd.get('db_password'):
        config_with_masked_pwd['db_password'] = '******'
    return JsonResponse({'success': True, 'config': config_with_masked_pwd})


@login_required
def save_backup_config_api(request):
    """API: 保存备份配置"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'})
    
    try:
        data = json.loads(request.body)
        
        # 获取现有配置作为基础
        current_config = get_backup_config()
        
        # 更新配置
        if 'db_host' in data:
            current_config['db_host'] = data['db_host']
        if 'db_port' in data:
            current_config['db_port'] = data['db_port']
        if 'db_user' in data:
            current_config['db_user'] = data['db_user']
        # 只有密码不是******时才更新
        if 'db_password' in data and data['db_password'] != '******':
            current_config['db_password'] = data['db_password']
        if 'db_name' in data:
            current_config['db_name'] = data['db_name']
        if 'backup_dir' in data:
            current_config['backup_dir'] = data['backup_dir']
        if 'keep_count' in data:
            current_config['keep_count'] = int(data['keep_count'])
        if 'auto_backup_enabled' in data:
            current_config['auto_backup_enabled'] = data['auto_backup_enabled']
        if 'auto_backup_time' in data:
            current_config['auto_backup_time'] = data['auto_backup_time']
            # 将时间转换为 cron 表达式
            time_parts = data['auto_backup_time'].split(':')
            if len(time_parts) == 2:
                hour = time_parts[0].strip()
                minute = time_parts[1].strip()
                current_config['auto_backup_cron'] = f'{minute} {hour} * * *'
        if 'auto_backup_cron' in data:
            current_config['auto_backup_cron'] = data['auto_backup_cron']
        if 'backup_media_enabled' in data:
            current_config['backup_media_enabled'] = data['backup_media_enabled']
        
        success, error = save_backup_config(current_config)
        
        if success:
            update_scheduler_job()
            return JsonResponse({'success': True, 'message': '备份配置保存成功'})
        else:
            return JsonResponse({'success': False, 'error': error})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def backup_list(request):
    return render(request, 'cmdb/backup_list.html')


@login_required
def create_database_backup(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'})
    
    try:
        config = get_backup_config()
        backup_dir = config['backup_dir']
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'cmdb_db_backup_{timestamp}.sql.gz'
        filepath = os.path.join(backup_dir, filename)
        
        cmd = [
            'mysqldump',
            '-h', config['db_host'],
            '-P', str(config['db_port']),
            '-u', config['db_user'],
            f"-p{config['db_password']}",
            '--ssl=0',
            '--single-transaction',
            '--quick',
            '--lock-tables=false',
            config['db_name']
        ]
        
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
        if result.returncode != 0:
            raise Exception(f"备份失败: {result.stderr.decode('utf-8')}")

        with gzip.open(filepath, 'wb') as f:
            f.write(result.stdout)

        file_size = os.path.getsize(filepath)
        
        media_backup_filename = None
        media_backup_size = 0
        media_backup_error = None
        
        if config.get('backup_media_enabled', True):
            try:
                media_backup_filename, media_backup_size = create_media_backup(backup_dir, timestamp)
            except Exception as e:
                media_backup_error = str(e)
        
        cleanup_old_backups(backup_dir, config.get('keep_count', 7), filename)
        
        size_str = f"{file_size / (1024 * 1024):.2f} MB" if file_size > 1024*1024 else f"{file_size / 1024:.2f} KB"
        BackupRecord.objects.create(
            backup_type='full',
            backup_name=filename,
            backup_path=filepath,
            file_size=size_str,
            status='success'
        )
        
        response_data = {
            'success': True,
            'filename': filename,
            'file_size': file_size,
            'message': '数据库备份成功'
        }
        
        if media_backup_filename:
            response_data['media_filename'] = media_backup_filename
            response_data['media_file_size'] = media_backup_size
            response_data['message'] += f'，媒体文件备份成功 ({format_file_size(media_backup_size)})'
        elif media_backup_error:
            response_data['media_error'] = media_backup_error
            response_data['message'] += f'，媒体文件备份失败: {media_backup_error}'
        
        return JsonResponse(response_data)
    except Exception as e:
        BackupRecord.objects.create(
            backup_type='full',
            backup_name=filename if 'filename' in dir() else 'unknown',
            status='failed'
        )
        return JsonResponse({'success': False, 'error': str(e)})


def create_media_backup(backup_dir, timestamp):
    """创建媒体文件备份"""
    from django.conf import settings
    
    media_root = settings.MEDIA_ROOT
    if not os.path.exists(media_root):
        return None, 0
    
    has_files = False
    for root, dirs, files in os.walk(media_root):
        if files:
            has_files = True
            break
    
    if not has_files:
        return None, 0
    
    media_filename = f'cmdb_media_backup_{timestamp}.tar.gz'
    media_filepath = os.path.join(backup_dir, media_filename)
    
    with tarfile.open(media_filepath, 'w:gz') as tar:
        tar.add(media_root, arcname='media')
    
    media_size = os.path.getsize(media_filepath)
    return media_filename, media_size


def format_file_size(size):
    """格式化文件大小"""
    if size > 1024 * 1024 * 1024:
        return f"{size / (1024 * 1024 * 1024):.2f} GB"
    elif size > 1024 * 1024:
        return f"{size / (1024 * 1024):.2f} MB"
    elif size > 1024:
        return f"{size / 1024:.2f} KB"
    else:
        return f"{size} B"


@login_required
def restore_database_backup(request, filename):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'})
    
    try:
        config = get_backup_config()
        backup_dir = config['backup_dir']
        filepath = os.path.join(backup_dir, filename)
        
        if not os.path.exists(filepath):
            return JsonResponse({'success': False, 'error': '备份文件不存在'})
        
        # 读取备份文件
        with gzip.open(filepath, 'rb') as f:
            sql_content = f.read()
        
        # 重要：在删除表之前刷新session，避免SessionInterrupted错误
        request.session.flush()
        
        # 步骤1: 删除所有现有表（避免类型不匹配问题）
        show_tables_cmd = [
            'mysql', '-h', config['db_host'], '-P', str(config['db_port']),
            '-u', config['db_user'], f"-p{config['db_password']}", '--ssl=0',
            '-s', '-r', '-e', f"SELECT table_name FROM information_schema.tables WHERE table_schema = '{config['db_name']}'",
            config['db_name']
        ]
        tables_result = subprocess.run(show_tables_cmd, capture_output=True, text=True)
        if tables_result.returncode == 0 and tables_result.stdout.strip():
            tables = tables_result.stdout.strip().split('\n')
            if tables:
                drop_sql = "SET FOREIGN_KEY_CHECKS=0;\n" + "\n".join([f"DROP TABLE IF EXISTS `{t}`;" for t in tables]) + "\nSET FOREIGN_KEY_CHECKS=1;"
                drop_cmd = [
                    'mysql', '-h', config['db_host'], '-P', str(config['db_port']),
                    '-u', config['db_user'], f"-p{config['db_password']}", '--ssl=0',
                    config['db_name']
                ]
                subprocess.run(drop_cmd, input=drop_sql.encode('utf-8'), capture_output=True)
        
        # 步骤2: 准备完整的SQL内容
        full_sql = b"SET FOREIGN_KEY_CHECKS=0;\n" + sql_content + b"\nSET FOREIGN_KEY_CHECKS=1;\n"
        
        # 步骤3: 执行恢复
        cmd = [
            'mysql',
            '-h', config['db_host'],
            '-P', str(config['db_port']),
            '-u', config['db_user'],
            f"-p{config['db_password']}",
            '--ssl=0',
            '--force',
            config['db_name']
        ]
        
        result = subprocess.run(cmd, input=full_sql, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
        if result.returncode != 0:
            error_msg = result.stderr.decode('utf-8', errors='replace')
            if "ERROR" in error_msg:
                error_lines = [line for line in error_msg.split('\n') if "ERROR" in line]
                if error_lines:
                    raise Exception(f"恢复失败: {error_lines[-1]}")
            raise Exception(f"恢复失败: {error_msg}")
        
        # 恢复成功后，返回简单HTML页面，避免session操作
        return HttpResponse('''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>数据恢复成功</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            margin: 0;
            padding: 20px;
        }
        .card {
            background: white;
            border-radius: 16px;
            padding: 40px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.2);
            text-align: center;
            max-width: 500px;
        }
        .icon {
            font-size: 64px;
            margin-bottom: 20px;
        }
        h1 {
            color: #155724;
            margin: 0 0 15px 0;
            font-size: 24px;
        }
        p {
            color: #555;
            line-height: 1.6;
            margin: 10px 0;
        }
        .button {
            display: inline-block;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 12px 30px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: 600;
            margin-top: 20px;
            transition: transform 0.2s;
        }
        .button:hover {
            transform: translateY(-2px);
        }
        .hint {
            background: #fff3cd;
            border: 1px solid #ffeeba;
            border-radius: 8px;
            padding: 15px;
            margin-top: 25px;
            color: #856404;
            font-size: 14px;
            text-align: left;
        }
    </style>
</head>
<body>
    <div class="card">
        <div class="icon">✓</div>
        <h1>数据库恢复成功！</h1>
        <p>您的数据已成功从备份文件恢复。</p>
        <div class="hint">
            <strong>💡 提示：</strong><br>
            由于恢复过程会重置session，请点击下方按钮重新登录系统。
        </div>
        <a href="/cmdb/login/" class="button">🔄 重新登录</a>
    </div>
</body>
</html>
''')
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@superuser_required
def database_restore(request):
    """数据恢复页面 - 支持上传SQL文件进行恢复"""
    restore_status = None
    
    if request.method == 'POST':
        if 'sql_file' not in request.FILES:
            restore_status = {'success': False, 'message': '请选择要上传的SQL文件'}
        else:
            sql_file = request.FILES['sql_file']
            
            # 验证文件类型
            if not sql_file.name.endswith('.sql') and not sql_file.name.endswith('.sql.gz'):
                restore_status = {'success': False, 'message': '只支持 .sql 或 .sql.gz 文件'}
            else:
                try:
                    import logging
                    logger = logging.getLogger(__name__)
                    
                    config = get_backup_config()
                    logger.info(f"开始数据恢复，文件: {sql_file.name}")
                    
                    # 读取文件内容
                    if sql_file.name.endswith('.sql.gz'):
                        import gzip
                        logger.info("解压SQL文件...")
                        sql_content = gzip.decompress(sql_file.read())
                    else:
                        sql_content = sql_file.read()
                    
                    logger.info(f"SQL文件大小: {len(sql_content)} 字节")
                    
                    # 重要：在删除表之前刷新session，避免SessionInterrupted错误
                    request.session.flush()
                    
                    # 步骤1: 删除所有现有表（避免类型不匹配问题）
                    logger.info("步骤1: 删除现有表...")
                    # 使用更可靠的方式删除表
                    show_tables_cmd = [
                        'mysql', '-h', config['db_host'], '-P', str(config['db_port']),
                        '-u', config['db_user'], f"-p{config['db_password']}", '--ssl=0',
                        '-s', '-r', '-e', f"SELECT table_name FROM information_schema.tables WHERE table_schema = '{config['db_name']}'",
                        config['db_name']
                    ]
                    tables_result = subprocess.run(show_tables_cmd, capture_output=True, text=True)
                    if tables_result.returncode == 0 and tables_result.stdout.strip():
                        tables = tables_result.stdout.strip().split('\n')
                        if tables:
                            drop_sql = "SET FOREIGN_KEY_CHECKS=0;\n" + "\n".join([f"DROP TABLE IF EXISTS `{t}`;" for t in tables]) + "\nSET FOREIGN_KEY_CHECKS=1;"
                            drop_cmd = [
                                'mysql', '-h', config['db_host'], '-P', str(config['db_port']),
                                '-u', config['db_user'], f"-p{config['db_password']}", '--ssl=0',
                                config['db_name']
                            ]
                            subprocess.run(drop_cmd, input=drop_sql.encode('utf-8'), capture_output=True)
                    logger.info("步骤1完成: 删除现有表成功")
                    
                    # 步骤2: 创建完整的SQL内容：先禁用外键检查，然后执行恢复，最后启用外键检查
                    logger.info("步骤2: 准备执行SQL恢复...")
                    full_sql = b"SET FOREIGN_KEY_CHECKS=0;\n" + sql_content + b"\nSET FOREIGN_KEY_CHECKS=1;\n"
                    
                    # 步骤3: 执行恢复（使用--force参数忽略错误）
                    logger.info("步骤3: 执行数据库恢复...")
                    cmd = [
                        'mysql',
                        '-h', config['db_host'],
                        '-P', str(config['db_port']),
                        '-u', config['db_user'],
                        f"-p{config['db_password']}",
                        '--ssl=0',
                        '--force',  # 强制继续执行，忽略错误
                        config['db_name']
                    ]
                    
                    # 使用Popen + communicate确保在同一连接中执行所有SQL
                    process = subprocess.Popen(cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    stdout, stderr = process.communicate(input=full_sql)
                    
                    if process.returncode != 0:
                        error_msg = stderr.decode('utf-8', errors='replace')
                        logger.error(f"数据库恢复失败: {error_msg}")
                        # 提取关键错误信息
                        if "ERROR" in error_msg:
                            error_lines = [line for line in error_msg.split('\n') if "ERROR" in line]
                            if error_lines:
                                raise Exception(f"数据库恢复失败: {error_lines[-1]}")
                        raise Exception(f"数据库恢复失败: {error_msg}")
                    
                    logger.info("步骤3完成: 数据库恢复成功")
                    # 恢复成功后，返回一个简单的HTML页面，避免session操作
                    return HttpResponse('''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>数据恢复成功</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            margin: 0;
            padding: 20px;
        }
        .card {
            background: white;
            border-radius: 16px;
            padding: 40px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.2);
            text-align: center;
            max-width: 500px;
        }
        .icon {
            font-size: 64px;
            margin-bottom: 20px;
        }
        h1 {
            color: #155724;
            margin: 0 0 15px 0;
            font-size: 24px;
        }
        p {
            color: #555;
            line-height: 1.6;
            margin: 10px 0;
        }
        .button {
            display: inline-block;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 12px 30px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: 600;
            margin-top: 20px;
            transition: transform 0.2s;
        }
        .button:hover {
            transform: translateY(-2px);
        }
        .hint {
            background: #fff3cd;
            border: 1px solid #ffeeba;
            border-radius: 8px;
            padding: 15px;
            margin-top: 25px;
            color: #856404;
            font-size: 14px;
            text-align: left;
        }
    </style>
</head>
<body>
    <div class="card">
        <div class="icon">✓</div>
        <h1>数据库恢复成功！</h1>
        <p>您的数据已成功从备份文件恢复。</p>
        <div class="hint">
            <strong>💡 提示：</strong><br>
            由于恢复过程会重置session，请点击下方按钮重新登录系统。
        </div>
        <a href="/cmdb/login/" class="button">🔄 重新登录</a>
    </div>
</body>
</html>
''')
                    
                except subprocess.TimeoutExpired:
                    restore_status = {'success': False, 'message': '恢复超时，请检查SQL文件大小或联系管理员'}
                except Exception as e:
                    import traceback
                    error_details = traceback.format_exc()
                    restore_status = {'success': False, 'message': f'恢复失败: {str(e)}'}

    return render(request, 'cmdb/database_restore.html', {'restore_status': restore_status})


@login_required
def delete_backup(request, filename):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'})

    try:
        config = get_backup_config()
        backup_dir = config['backup_dir']
        filepath = os.path.join(backup_dir, filename)

        if not os.path.exists(filepath):
            return JsonResponse({'success': False, 'error': '文件不存在'})

        os.remove(filepath)

        log_operation(request.user, 'delete', f'备份文件: {filename}', f'删除备份文件: {filename}', request.META.get('REMOTE_ADDR'))

        return JsonResponse({'success': True, 'message': '备份文件删除成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def download_backup(request, filename):
    config = get_backup_config()
    backup_dir = config['backup_dir']
    filepath = os.path.join(backup_dir, filename)
    
    if not os.path.exists(filepath):
        return JsonResponse({'success': False, 'error': '文件不存在'})
    
    try:
        with open(filepath, 'rb') as f:
            content = f.read()
        
        from django.http import HttpResponse
        response = HttpResponse(content, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def restore_media_backup(request, filename):
    """恢复媒体文件备份"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'})
    
    if not filename.endswith('.tar.gz'):
        return JsonResponse({'success': False, 'error': '无效的媒体备份文件（只支持 .tar.gz 格式）'})
    
    try:
        config = get_backup_config()
        backup_dir = config['backup_dir']
        filepath = os.path.join(backup_dir, filename)
        
        if not os.path.exists(filepath):
            return JsonResponse({'success': False, 'error': '备份文件不存在'})
        
        media_root = settings.MEDIA_ROOT
        if os.path.exists(media_root):
            backup_media_dir = os.path.join(media_root, f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
            os.makedirs(backup_media_dir, exist_ok=True)
            for item in os.listdir(media_root):
                if item.startswith('backup_'):
                    continue
                item_path = os.path.join(media_root, item)
                if os.path.isdir(item_path):
                    shutil.move(item_path, os.path.join(backup_media_dir, item))
                else:
                    shutil.move(item_path, os.path.join(backup_media_dir, item))
        
        with tarfile.open(filepath, 'r:gz') as tar:
            tar.extractall(path=os.path.dirname(media_root))
        
        return JsonResponse({
            'success': True,
            'message': f'媒体文件恢复成功！原文件已备份到 {backup_media_dir if "backup_media_dir" in dir() else "无"}'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def restore_media_backup_upload(request):
    """通过上传文件恢复媒体备份"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'})
    
    try:
        if 'media_file' not in request.FILES:
            return JsonResponse({'success': False, 'error': '请选择要上传的媒体备份文件'})
        
        media_file = request.FILES['media_file']
        filename = media_file.name
        
        # 检查文件扩展名，支持多种情况
        if not (filename.endswith('.tar.gz') or filename.endswith('.tgz')):
            return JsonResponse({'success': False, 'error': f'只支持 .tar.gz 格式的备份文件（当前文件：{filename}）'})
        
        media_root = settings.MEDIA_ROOT
        backup_media_dir = None
        
        if os.path.exists(media_root):
            backup_media_dir = os.path.join(media_root, f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
            os.makedirs(backup_media_dir, exist_ok=True)
            for item in os.listdir(media_root):
                if item.startswith('backup_'):
                    continue
                item_path = os.path.join(media_root, item)
                if os.path.isdir(item_path):
                    shutil.move(item_path, os.path.join(backup_media_dir, item))
                else:
                    shutil.move(item_path, os.path.join(backup_media_dir, item))
        
        # 将上传的文件保存到临时位置再解压
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.tar.gz', delete=False) as temp_file:
            temp_path = temp_file.name
            for chunk in media_file.chunks():
                temp_file.write(chunk)
        
        try:
            with tarfile.open(temp_path, 'r:gz') as tar:
                # 检查是否是有效的tar文件
                members = tar.getmembers()
                if len(members) == 0:
                    return JsonResponse({'success': False, 'error': '无效的媒体备份文件（文件为空）'})
                
                # 修复权限问题：重置所有文件的权限，避免root权限文件无法访问
                for member in members:
                    # 重置权限为可读写
                    member.mode = 0o644 if member.isfile() else 0o755
                
                tar.extractall(path=os.path.dirname(media_root))
        except tarfile.ReadError as e:
            return JsonResponse({'success': False, 'error': f'无效的媒体备份文件：{str(e)}'})
        except tarfile.CompressionError as e:
            return JsonResponse({'success': False, 'error': f'文件压缩格式错误：{str(e)}'})
        except PermissionError as e:
            return JsonResponse({'success': False, 'error': f'权限不足：{str(e)}。请检查媒体目录权限'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'解压文件失败：{str(e)}'})
        finally:
            # 清理临时文件
            if os.path.exists(temp_path):
                os.remove(temp_path)
        
        message = '媒体文件恢复成功！'
        if backup_media_dir:
            message += f' 原文件已备份到 {backup_media_dir}'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_backup_list_with_stats(request):
    config = get_backup_config()
    backup_dir = config['backup_dir']
    files = []
    total_size = 0
    
    if os.path.exists(backup_dir):
        file_list = sorted(os.listdir(backup_dir), reverse=True)
        for fname in file_list:
            filepath = os.path.join(backup_dir, fname)
            
            # 支持目录格式的备份
            if os.path.isdir(filepath) and fname.startswith('cmdb_project_'):
                # 计算目录大小
                dir_size = 0
                for root, dirs, filenames in os.walk(filepath):
                    for f in filenames:
                        dir_size += os.path.getsize(os.path.join(root, f))
                total_size += dir_size
                
                try:
                    # 从目录名提取时间戳
                    time_str = fname.replace('cmdb_project_', '')
                    created_at = datetime.strptime(time_str, '%Y%m%d%H%M')
                    created_at_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    created_at_str = '未知'
                
                backup_type = 'manual'
                
                if dir_size < 1024:
                    size_str = f"{dir_size} B"
                elif dir_size < 1024 * 1024:
                    size_str = f"{dir_size / 1024:.2f} KB"
                else:
                    size_str = f"{dir_size / (1024 * 1024):.2f} MB"
                
                files.append({
                    'filename': fname,
                    'created_at': created_at_str,
                    'size': size_str,
                    'file_size': dir_size,
                    'backup_type': backup_type,
                    'is_directory': True
                })
            
            # 支持 .sql.gz 和 .json.gz 文件格式的备份
            elif (fname.endswith('.sql.gz') or fname.endswith('.json.gz')) and os.path.isfile(filepath):
                file_size = os.path.getsize(filepath)
                total_size += file_size
                
                try:
                    time_str = fname.replace('cmdb_db_backup_', '').replace('.sql.gz', '')
                    created_at = datetime.strptime(time_str, '%Y%m%d_%H%M%S')
                    created_at_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    created_at_str = '未知'
                
                backup_type = 'manual' if '_manual_' in fname else 'auto'
                
                if file_size < 1024:
                    size_str = f"{file_size} B"
                elif file_size < 1024 * 1024:
                    size_str = f"{file_size / 1024:.2f} KB"
                else:
                    size_str = f"{file_size / (1024 * 1024):.2f} MB"
                
                files.append({
                    'filename': fname,
                    'created_at': created_at_str,
                    'size': size_str,
                    'file_size': file_size,
                    'backup_type': backup_type,
                    'is_directory': False
                })
            
            # 支持媒体备份文件 .tar.gz
            elif fname.endswith('.tar.gz') and fname.startswith('cmdb_media_backup_') and os.path.isfile(filepath):
                file_size = os.path.getsize(filepath)
                total_size += file_size
                
                try:
                    time_str = fname.replace('cmdb_media_backup_', '').replace('.tar.gz', '')
                    created_at = datetime.strptime(time_str, '%Y%m%d_%H%M%S')
                    created_at_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    created_at_str = '未知'
                
                if file_size < 1024:
                    size_str = f"{file_size} B"
                elif file_size < 1024 * 1024:
                    size_str = f"{file_size / 1024:.2f} KB"
                else:
                    size_str = f"{file_size / (1024 * 1024):.2f} MB"
                
                files.append({
                    'filename': fname,
                    'created_at': created_at_str,
                    'size': size_str,
                    'file_size': file_size,
                    'backup_type': 'media',
                    'is_directory': False
                })
    
    try:
        disk = shutil.disk_usage(backup_dir)
        disk_free = disk.free
    except:
        disk_free = 0
    
    return JsonResponse({
        'files': files,
        'stats': {
            'total_files': len(files),
            'total_size': total_size,
            'disk_free': disk_free
        }
    })


def cleanup_old_backups(backup_dir, keep_count, current_file=None):
    try:
        db_files = []
        media_files = []
        for fname in os.listdir(backup_dir):
            filepath = os.path.join(backup_dir, fname)
            if os.path.isfile(filepath):
                if fname.endswith('.sql.gz'):
                    db_files.append({
                        'name': fname, 'path': filepath, 'mtime': os.path.getmtime(filepath)
                    })
                elif fname.endswith('.tar.gz') and fname.startswith('cmdb_media_backup_'):
                    media_files.append({
                        'name': fname, 'path': filepath, 'mtime': os.path.getmtime(filepath)
                    })

        db_files.sort(key=lambda x: x['mtime'], reverse=True)
        media_files.sort(key=lambda x: x['mtime'], reverse=True)

        for i, f in enumerate(db_files):
            if i >= keep_count:
                try:
                    os.remove(f['path'])
                except:
                    pass
        
        for i, f in enumerate(media_files):
            if i >= keep_count:
                try:
                    os.remove(f['path'])
                except:
                    pass
    except Exception as e:
        print(f"清理旧备份失败: {str(e)}")


def user_login(request):
    if request.user.is_authenticated:
        return redirect('index')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        next_url = request.POST.get('next', '/cmdb/')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_active:
                login(request, user)
                messages.success(request, f'欢迎回来，{user.username}！')
                log_operation(user, "login", target=user.username, description="用户登录")
                return redirect(next_url)
            else:
                messages.error(request, '该账户已被禁用，请联系管理员。')
        else:
            messages.error(request, '用户名或密码错误。')

    return render(request, 'registration/login.html', {'next': request.GET.get('next', '/cmdb/')})


@login_required
def user_logout(request):
    logout(request)
    log_operation(request.user, "logout", description="用户登出")
    messages.success(request, '您已成功登出。')
    return redirect('login')


@login_required
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # 验证新密码是否一致
        if new_password != confirm_password:
            messages.error(request, '两次输入的新密码不一致')
            return redirect('change_password')
        
        # 验证旧密码是否正确
        if not request.user.check_password(old_password):
            messages.error(request, '旧密码不正确')
            return redirect('change_password')
        
        # 修改密码
        request.user.set_password(new_password)
        request.user.save()
        messages.success(request, '密码修改成功，请重新登录')
        return redirect('login')
    
    return render(request, 'cmdb/change_password.html')


@login_required
def user_guide(request):
    return render(request, 'cmdb/user_guide.html')


@login_required
@superuser_required
def user_management(request):
    users = User.objects.select_related('userprofile').all().order_by('-id')
    
    # 确保每个用户都有 UserProfile
    for user in users:
        UserProfile.objects.get_or_create(user=user)
    
    # 重新获取用户列表
    users = User.objects.select_related('userprofile').all().order_by('-id')

    return render(request, 'cmdb/settings.html', {
        'users': users,
        'active_tab': 'users'
    })


@login_required
@admin_required
def user_add(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email', '')
        real_name = request.POST.get('real_name', '')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        is_active = request.POST.get('is_active', '1') == '1'
        is_admin = request.POST.get('is_admin') == 'on'

        if password != confirm_password:
            messages.error(request, '两次输入的密码不一致。')
            return redirect('user_add')

        if User.objects.filter(username=username).exists():
            messages.error(request, '用户名已存在。')
            return redirect('user_add')

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                is_active=is_active
            )

            UserProfile.objects.create(
                user=user,
                real_name=real_name,
                is_admin=is_admin
            )

            messages.success(request, f'用户 {username} 创建成功！')
            return redirect('user_management')
        except Exception as e:
            messages.error(request, f'创建用户失败：{str(e)}')
            return redirect('user_add')

    return render(request, 'cmdb/user_form.html', {
        'user': None,
        'profile': None,
        'is_admin': False
    })


@login_required
@admin_required
def user_edit(request, user_id):
    edit_user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        email = request.POST.get('email', '')
        real_name = request.POST.get('real_name', '')
        password = request.POST.get('password')
        is_active = request.POST.get('is_active', '1') == '1'
        is_admin = request.POST.get('is_admin') == 'on'

        # 检查是否要取消管理员权限，确保至少有一个管理员
        profile = UserProfile.objects.get_or_create(user=edit_user)[0]
        current_is_admin = profile.is_admin
        
        if current_is_admin and not is_admin:
            # 检查是否还有其他管理员
            admin_count = UserProfile.objects.filter(is_admin=True).count()
            if admin_count <= 1 and not edit_user.is_superuser:
                messages.error(request, '系统必须至少有一个管理员用户！')
                return redirect('user_edit', user_id=user_id)

        try:
            edit_user.email = email
            edit_user.is_active = is_active
            edit_user.save()

            profile.real_name = real_name
            profile.is_admin = is_admin
            profile.save()

            if password:
                edit_user.set_password(password)
                edit_user.save()

            messages.success(request, f'用户 {edit_user.username} 信息已更新！')
            return redirect('user_management')
        except Exception as e:
            messages.error(request, f'更新用户失败：{str(e)}')

    profile = UserProfile.objects.get_or_create(user=edit_user)[0]
    
    return render(request, 'cmdb/user_form.html', {
        'user': edit_user,
        'profile': profile,
        'is_admin': profile.is_admin
    })


@login_required
@admin_required
def user_permissions(request, user_id):
    # 权限系统已简化，管理员和普通用户的区别仅在于是否能访问系统设置
    # 此视图保留但重定向到用户编辑页面
    return redirect('user_edit', user_id=user_id)


@login_required
@admin_required
def user_disable(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.user.id == user_id:
        messages.error(request, '不能禁用自己的账户！')
        return redirect('user_management')

    # 检查是否是最后一个管理员
    try:
        profile = user.userprofile
        if profile.is_admin:
            admin_count = UserProfile.objects.filter(is_admin=True).count()
            if admin_count <= 1 and not user.is_superuser:
                messages.error(request, '系统必须至少有一个管理员用户！')
                return redirect('user_management')
    except UserProfile.DoesNotExist:
        pass

    user.is_active = False
    user.save()
    messages.success(request, f'用户 {user.username} 已禁用。')
    return redirect('user_management')


@login_required
def user_enable(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.is_active = True
    user.save()
    messages.success(request, f'用户 {user.username} 已启用。')
    return redirect('user_management')


@login_required
@admin_required
def user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.user.id == user_id:
        messages.error(request, '不能删除自己的账户！')
        return redirect('user_management')

    # 检查是否是最后一个管理员
    try:
        profile = user.userprofile
        if profile.is_admin:
            admin_count = UserProfile.objects.filter(is_admin=True).count()
            if admin_count <= 1 and not user.is_superuser:
                messages.error(request, '系统必须至少有一个管理员用户！')
                return redirect('user_management')
    except UserProfile.DoesNotExist:
        pass

    username = user.username
    user.delete()
    messages.success(request, f'用户 {username} 已删除。')
    return redirect('user_management')


@login_required
@admin_required
def role_management(request):
    # 角色管理作为预留功能，当前权限系统简化为管理员/普通用户
    messages.info(request, '当前系统使用简化的权限模型，仅区分管理员和普通用户。角色管理功能作为预留功能保留。')
    return redirect('user_management')


@login_required
@admin_required
def permission_management(request):
    # 权限管理作为预留功能，当前权限系统简化为管理员/普通用户
    messages.info(request, '当前系统使用简化的权限模型，管理员可以访问系统设置，普通用户不能。权限管理功能作为预留功能保留。')
    return redirect('user_management')


# 服务器备件相关视图

@login_required
def spareparts_list(request):
    """备件列表页面"""
    # 只显示未安装的备件（库存中、维修中、空状态等）
    spareparts = SparePart.objects.filter(status__in=['in_stock', 'maintenance', 'available', '']).order_by('-created_at')
    static_assets = StaticAsset.objects.all().order_by('asset_no')
    part_types = SparePartType.objects.filter(is_active=True).order_by('order')
    return render(request, 'cmdb/spareparts/list.html', {
        'spareparts': spareparts,
        'static_assets': static_assets,
        'part_types': part_types
    })


@login_required
def server_spareparts_list(request):
    """服务器备件列表页面"""
    spareparts = SparePart.objects.filter(category='server', status__in=['in_stock', 'maintenance', 'available', '']).order_by('-created_at')
    static_assets = StaticAsset.objects.all().order_by('asset_no')
    part_types = SparePartType.objects.filter(is_active=True).order_by('order')
    return render(request, 'cmdb/spareparts/list.html', {
        'spareparts': spareparts,
        'static_assets': static_assets,
        'part_types': part_types,
        'category': 'server',
        'category_name': '服务器备件'
    })


@login_required
def desktop_spareparts_list(request):
    """办公机备件列表页面"""
    spareparts = SparePart.objects.filter(category='desktop', status__in=['in_stock', 'maintenance', 'available', '']).order_by('-created_at')
    static_assets = StaticAsset.objects.all().order_by('asset_no')
    part_types = SparePartType.objects.filter(is_active=True).order_by('order')
    return render(request, 'cmdb/spareparts/list.html', {
        'spareparts': spareparts,
        'static_assets': static_assets,
        'part_types': part_types,
        'category': 'desktop',
        'category_name': '办公机备件'
    })


@login_required
def spareparts_add(request):
    """添加备件"""
    if request.method == 'POST':
        try:
            sparepart = SparePart()
            sparepart.asset_code = request.POST.get('asset_code') or None
            sparepart.name = request.POST.get('name')
            sparepart.brand = request.POST.get('brand')
            sparepart.model = request.POST.get('model')
            sparepart.size = request.POST.get('size')
            sparepart.serial_number = request.POST.get('serial_number')
            sparepart.location = request.POST.get('location')
            sparepart.purchase_date = request.POST.get('purchase_date') or None
            
            type_id = request.POST.get('type')
            if type_id:
                sparepart.type_id = type_id
            
            # 处理图片上传 - 按年月日创建目录
            images = []
            if request.FILES:
                import datetime
                date_dir = datetime.datetime.now().strftime('%Y%m%d')
                for key in request.FILES:
                    file = request.FILES[key]
                    import uuid
                    ext = file.name.split('.')[-1]
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    filepath = os.path.join(settings.MEDIA_ROOT, 'spareparts', date_dir, filename)
                    os.makedirs(os.path.dirname(filepath), exist_ok=True)
                    with open(filepath, 'wb') as f:
                        for chunk in file.chunks():
                            f.write(chunk)
                    images.append(f'/media/spareparts/{date_dir}/{filename}')
            
            if images:
                sparepart.images = json.dumps(images)
            
            sparepart.save()
            
            log_operation(request.user, 'add', f'备件: {sparepart.name}', '添加备件', request.META.get('REMOTE_ADDR'))
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # 获取备件类型列表
    part_types = SparePartType.objects.filter(is_active=True).order_by('order')
    return render(request, 'cmdb/spareparts/modal_add.html', {'part_types': part_types})

@login_required
def spareparts_edit(request, sparepart_id):
    """编辑备件"""
    sparepart = get_object_or_404(SparePart, id=sparepart_id)
    
    if request.method == 'POST':
        try:
            sparepart.asset_code = request.POST.get('asset_code') or None
            sparepart.name = request.POST.get('name')
            sparepart.brand = request.POST.get('brand')
            sparepart.model = request.POST.get('model')
            sparepart.size = request.POST.get('size')
            sparepart.serial_number = request.POST.get('serial_number')
            sparepart.location = request.POST.get('location')
            sparepart.purchase_date = request.POST.get('purchase_date') or None
            
            type_id = request.POST.get('type')
            if type_id:
                sparepart.type_id = type_id
            else:
                sparepart.type_id = None
            
            # 处理图片上传 - 按年月日创建目录
            # 获取剩余的现有图片路径
            remaining_images = []
            if request.POST.get('remaining_images'):
                try:
                    remaining_images = json.loads(request.POST.get('remaining_images'))
                except:
                    pass
            
            # 处理新上传的图片
            new_image_paths = []
            if request.FILES.getlist('images'):
                images = request.FILES.getlist('images')
                import datetime
                date_dir = datetime.datetime.now().strftime('%Y%m%d')
                for img in images:
                    filename = f"{uuid.uuid4()}_{img.name}"
                    filepath = os.path.join(settings.MEDIA_ROOT, 'spareparts', date_dir, filename)
                    os.makedirs(os.path.dirname(filepath), exist_ok=True)
                    with open(filepath, 'wb+') as destination:
                        for chunk in img.chunks():
                            destination.write(chunk)
                    new_image_paths.append(f"/media/spareparts/{date_dir}/{filename}")
            
            # 合并剩余图片和新上传的图片
            all_images = remaining_images + new_image_paths
            if all_images:
                sparepart.images = json.dumps(all_images)
            else:
                sparepart.images = ''
            
            sparepart.save()
            
            log_operation(request.user, 'update', f'备件: {sparepart.name}', '编辑备件', request.META.get('REMOTE_ADDR'))
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET请求返回JSON数据
    images_list = []
    if sparepart.images:
        try:
            images_list = json.loads(sparepart.images)
        except:
            pass
    
    return JsonResponse({
        'success': True,
        'data': {
            'id': sparepart.id,
            'asset_code': sparepart.asset_code,
            'name': sparepart.name,
            'brand': sparepart.brand,
            'model': sparepart.model,
            'size': sparepart.size,
            'serial_number': sparepart.serial_number,
            'location': sparepart.location,
            'purchase_date': sparepart.purchase_date.isoformat() if sparepart.purchase_date else None,
            'type_id': sparepart.type_id,
            'status': sparepart.status,
            'images': images_list
        }
    })

@login_required
def spareparts_delete(request, sparepart_id):
    """删除备件"""
    if request.method == 'POST':
        try:
            sparepart = get_object_or_404(SparePart, id=sparepart_id)
            target = f'备件: {sparepart.name}'
            sparepart.delete()
            
            log_operation(request.user, 'delete', target, '删除备件', request.META.get('REMOTE_ADDR'))
            return JsonResponse({'success': True, 'message': '删除成功'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})

# 备件类型管理
@login_required
def sparepart_types_list(request):
    """备件类型列表"""
    types = SparePartType.objects.all().order_by('order')
    return render(request, 'cmdb/spareparts/types_list.html', {'types': types})

@login_required
def sparepart_type_add(request):
    """添加备件类型"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            code = request.POST.get('code')
            
            if SparePartType.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'error': '类型代码已存在'})
            
            SparePartType.objects.create(name=name, code=code, order=request.POST.get('order', 0))
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return render(request, 'cmdb/spareparts/modal_type_add.html')

@login_required
def sparepart_type_edit(request, type_id):
    """编辑备件类型"""
    part_type = get_object_or_404(SparePartType, id=type_id)
    
    if request.method == 'POST':
        try:
            part_type.name = request.POST.get('name')
            old_code = part_type.code
            new_code = request.POST.get('code')
            
            if old_code != new_code and SparePartType.objects.filter(code=new_code).exists():
                return JsonResponse({'success': False, 'error': '类型代码已存在'})
            
            part_type.code = new_code
            part_type.order = request.POST.get('order', 0)
            part_type.is_active = request.POST.get('is_active') == 'on'
            part_type.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET请求返回JSON数据
    return JsonResponse({
        'id': part_type.id,
        'name': part_type.name,
        'code': part_type.code,
        'order': part_type.order,
        'is_active': part_type.is_active
    })

@login_required
def sparepart_type_delete(request, type_id):
    """删除备件类型"""
    if request.method == 'POST':
        try:
            part_type = get_object_or_404(SparePartType, id=type_id)
            part_type.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})

@login_required
def api_get_assets(request):
    """获取可用资产列表API（用于下拉选择）"""
    assets = StaticAsset.objects.filter(asset_no__isnull=False).values('asset_no', 'ip', 'device_model', 'cabinet')
    assets_list = list(assets)
    return JsonResponse({'assets': assets_list})

@login_required
def api_get_asset_info(request, asset_no):
    """获取单个资产信息API"""
    try:
        asset = StaticAsset.objects.get(asset_no=asset_no)
        return JsonResponse({
            'asset_no': asset.asset_no,
            'ip': str(asset.ip) if asset.ip else '',
            'cabinet': asset.cabinet,
            'device_model': asset.device_model
        })
    except StaticAsset.DoesNotExist:
        return JsonResponse({'error': '资产不存在'}, status=404)

# 测试侧边栏页面
def test_sidebar(request):
    return render(request, 'test_sidebar.html')


# ==================== 资产关系管理 ====================

@login_required
def asset_relation_list(request):
    """资产关系列表页面"""
    search_query = request.GET.get('search', '')
    relations = AssetRelation.objects.filter(is_active=True).select_related('parent_asset', 'child_asset')
    
    # 搜索过滤
    if search_query:
        relations = relations.filter(
            Q(parent_asset__asset_no__icontains=search_query) |
            Q(parent_asset__ip__icontains=search_query) |
            Q(child_asset__asset_no__icontains=search_query) |
            Q(child_asset__hostname__icontains=search_query) |
            Q(child_asset__device_model__icontains=search_query) |
            Q(child_asset__sn__icontains=search_query) |
            Q(slot__icontains=search_query)
        )
    
    static_assets = StaticAsset.objects.all().order_by('asset_no')
    spareparts = SparePart.objects.filter(status='in_stock').order_by('-created_at')
    return render(request, 'cmdb/asset_relation/list.html', {
        'relations': relations,
        'static_assets': static_assets,
        'spareparts': spareparts,
        'search_query': search_query
    })


@login_required
def asset_relation_detail(request, relation_id):
    """资产关系详情"""
    relation = get_object_or_404(AssetRelation, id=relation_id)
    histories = InstallHistory.objects.filter(asset_relation=relation).order_by('-install_time')
    return render(request, 'cmdb/asset_relation/detail.html', {
        'relation': relation,
        'histories': histories
    })


@login_required
def api_get_host_children(request, host_id):
    """获取主机的子资产列表"""
    relations = AssetRelation.objects.filter(parent_asset_id=host_id, is_active=True)\
        .select_related('child_asset')
    children = []
    for rel in relations:
        children.append({
            'id': rel.child_asset.id,
            'name': rel.child_asset.name,
            'model': rel.child_asset.model,
            'serial_number': rel.child_asset.serial_number,
            'slot': rel.slot,
            'is_removable': rel.is_removable,
            'relation_id': rel.id
        })
    return JsonResponse({'children': children})


@login_required
def api_install_component(request):
    """安装已有组件到主资产"""
    if request.method == 'POST':
        try:
            data = request.POST
            parent_host_id = data.get('parent_host_id')
            child_host_id = data.get('child_host_id')
            slot = data.get('slot')
            remark = data.get('remark', '')

            with transaction.atomic():
                child_host = Host.objects.select_for_update().get(id=child_host_id)
                parent_host = Host.objects.select_for_update().get(id=parent_host_id)

                existing_relation = AssetRelation.objects.select_for_update()\
                    .filter(child_asset_id=child_host_id, is_active=True).first()
                if existing_relation:
                    return JsonResponse({'success': False, 'error': '该组件已安装在其他资产上'})

                slot_occupied = AssetRelation.objects.select_for_update()\
                    .filter(parent_asset_id=parent_host_id, slot=slot, is_active=True).exists()
                if slot_occupied:
                    return JsonResponse({'success': False, 'error': f'槽位 {slot} 已被占用'})

                relation = AssetRelation.objects.create(
                    parent_asset=parent_host,
                    child_asset=child_host,
                    slot=slot,
                    is_removable=True,
                    is_active=True
                )

                InstallHistory.objects.create(
                    asset_relation=relation,
                    parent_asset=parent_host,
                    child_asset=child_host,
                    install_time=timezone.now(),
                    operator=request.user,
                    operation_type='install',
                    source_type='direct_purchase',
                    remark=remark
                )

                LifecycleEvent.objects.create(
                    asset=child_host,
                    event_type='deploy',
                    event_time=timezone.now(),
                    operator=request.user,
                    remark=f'安装到主机 {parent_host.name} 的 {slot} 槽位'
                )

            return JsonResponse({'success': True, 'relation_id': relation.id})
        except Host.DoesNotExist:
            return JsonResponse({'success': False, 'error': '主机不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_uninstall_component(request):
    """拆卸组件"""
    if request.method == 'POST':
        try:
            data = request.POST
            relation_id = data.get('relation_id')
            action = data.get('action')
            remark = data.get('remark', '')
            sparepart_name = data.get('sparepart_name', '')

            with transaction.atomic():
                relation = AssetRelation.objects.select_for_update().select_related('parent_asset', 'child_asset').get(id=relation_id)
                child_host = relation.child_asset
                parent_host = relation.parent_asset

                relation.is_active = False
                relation.save()

                history = InstallHistory.objects.filter(asset_relation=relation, uninstall_time__isnull=True).first()
                if history:
                    history.uninstall_time = timezone.now()
                    history.operation_type = 'uninstall'
                    history.remark = remark
                    history.save()
                else:
                    InstallHistory.objects.create(
                        asset_relation=relation,
                        parent_asset=parent_host,
                        child_asset=child_host,
                        install_time=relation.created_at,
                        uninstall_time=timezone.now(),
                        operator=request.user,
                        operation_type='uninstall',
                        remark=remark
                    )

                if action == 'return_to_spare':
                    SparePart.objects.create(
                        name=sparepart_name or child_host.hostname,
                        model=child_host.device_model or '',
                        serial_number=child_host.sn or '',
                        status='in_stock',
                        is_installed=False,
                        installed_host_id=None,
                        installed_slot='',
                    )
                    LifecycleEvent.objects.create(
                        asset=child_host,
                        event_type='uninstall',
                        event_time=timezone.now(),
                        operator=request.user,
                        remark=f'拆卸返回备件库: {remark}'
                    )
                else:
                    LifecycleEvent.objects.create(
                        asset=child_host,
                        event_type='scrap',
                        event_time=timezone.now(),
                        operator=request.user,
                        remark=f'拆卸报废: {remark}'
                    )

            return JsonResponse({'success': True})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_get_install_history(request, relation_id):
    """获取安装历史"""
    try:
        histories = InstallHistory.objects.filter(asset_relation_id=relation_id)\
            .select_related('operator').order_by('-install_time')
        result = []
        for h in histories:
            result.append({
                'id': h.id,
                'install_time': h.install_time.strftime('%Y-%m-%d %H:%M:%S'),
                'uninstall_time': h.uninstall_time.strftime('%Y-%m-%d %H:%M:%S') if h.uninstall_time else '',
                'operator': h.operator.username if h.operator else '',
                'operation_type': h.get_operation_type_display(),
                'source_type': h.get_source_type_display() if h.source_type else '',
                'remark': h.remark
            })
        return JsonResponse({'histories': result})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==================== 备件安装 ====================

@login_required
def api_install_sparepart(request):
    """从备件库安装备件到资产"""
    if request.method == 'POST':
        try:
            data = request.POST
            sparepart_id = data.get('sparepart_id')
            static_asset_id = data.get('parent_host_id')  # 现在是StaticAsset的ID
            slot = data.get('slot')
            remark = data.get('remark', '')

            with transaction.atomic():
                sparepart = SparePart.objects.select_for_update().get(id=sparepart_id)
                # 允许库存中、维修中以及旧状态值的备件进行安装
                valid_statuses = ['in_stock', 'maintenance', 'available', '']
                if sparepart.status not in valid_statuses:
                    return JsonResponse({'success': False, 'error': f'备件当前状态为"{dict(SparePart.STATUS_CHOICES).get(sparepart.status, sparepart.status)}"，无法安装'})

                # 获取静态资产信息
                static_asset = StaticAsset.objects.select_for_update().get(id=static_asset_id)
                
                # 根据静态资产找到或创建对应的Host记录
                parent_host, created = Host.objects.get_or_create(
                    asset_no=static_asset.asset_no,
                    defaults={
                        'hostname': static_asset.asset_no or f"host-{static_asset.id}",
                        'ip': static_asset.ip or '0.0.0.0',
                        'device_model': static_asset.device_model,
                        'status': static_asset.status
                    }
                )

                slot_occupied = AssetRelation.objects.select_for_update()\
                    .filter(parent_asset_id=parent_host.id, slot=slot, is_active=True).exists()
                if slot_occupied:
                    return JsonResponse({'success': False, 'error': f'槽位 {slot} 已被占用'})

                hostname = f"component-{sparepart.id}-{sparepart.name.replace(' ', '-').lower()}-{int(timezone.now().timestamp())}"
                child_host, created = Host.objects.get_or_create(
                    hostname=hostname,
                    defaults={
                        'device_model': sparepart.model,
                        'sn': sparepart.serial_number,
                        'ip': '0.0.0.0',
                        'memo': sparepart.name,  # 设置显示名称为备件名称
                        'asset_no': sparepart.asset_code,  # 备件资产编号
                        'images': sparepart.images  # 备件图片
                    }
                )

                relation = AssetRelation.objects.create(
                    parent_asset=parent_host,
                    child_asset=child_host,
                    slot=slot,
                    is_removable=True,
                    is_active=True
                )

                InstallHistory.objects.create(
                    asset_relation=relation,
                    parent_asset=parent_host,
                    child_asset=child_host,
                    install_time=timezone.now(),
                    operator=request.user,
                    operation_type='install',
                    source_type='spare',
                    remark=f'从备件库安装: {sparepart.name}'
                )

                sparepart.status = 'installed'
                sparepart.is_installed = True
                sparepart.installed_host_id = parent_host.id
                sparepart.installed_slot = slot
                sparepart.save()

                LifecycleEvent.objects.create(
                    asset=child_host,
                    event_type='deploy',
                    event_time=timezone.now(),
                    operator=request.user,
                    remark=f'从备件库安装到主机 {parent_host.hostname} 的 {slot} 槽位'
                )

            return JsonResponse({'success': True, 'relation_id': relation.id})
        except SparePart.DoesNotExist:
            return JsonResponse({'success': False, 'error': '备件不存在'})
        except Host.DoesNotExist:
            return JsonResponse({'success': False, 'error': '主机不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_return_sparepart(request):
    """将备件退库到已退库设备列表"""
    if request.method == 'POST':
        try:
            data = request.POST
            sparepart_id = data.get('sparepart_id')
            return_reason = data.get('return_reason', '')

            with transaction.atomic():
                sparepart = SparePart.objects.select_for_update().get(id=sparepart_id)
                
                # 创建子资产Host记录
                hostname = f"returned-sparepart-{sparepart.id}-{int(timezone.now().timestamp())}"
                child_host = Host.objects.create(
                    hostname=hostname,
                    device_model=sparepart.model,
                    sn=sparepart.serial_number,
                    ip='0.0.0.0',
                    memo=sparepart.name,
                    asset_no=sparepart.asset_code,
                    images=sparepart.images
                )

                # 创建资产关系，标记为已退库
                relation = AssetRelation.objects.create(
                    parent_asset=child_host,
                    child_asset=child_host,
                    slot=sparepart.location,
                    is_removable=True,
                    is_active=False,
                    is_returned=True,
                    returned_at=timezone.now()
                )

                # 记录操作日志
                log_operation(request.user, 'update', f'备件退库: {sparepart.name}', '备件退库操作', request.META.get('REMOTE_ADDR'))

                # 删除备件记录
                sparepart.delete()

                return JsonResponse({'success': True, 'message': '备件退库成功'})
        except SparePart.DoesNotExist:
            return JsonResponse({'success': False, 'error': '备件不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_direct_install_component(request):
    """直接安装新组件（未经过备件库）"""
    if request.method == 'POST':
        try:
            data = request.POST
            static_asset_id = data.get('parent_host_id')  # 现在是StaticAsset的ID
            component_name = data.get('component_name')
            component_asset_no = data.get('component_asset_no', '')  # 新增字段
            component_model = data.get('component_model', '')
            component_sn = data.get('component_sn', '')
            slot = data.get('slot', '')
            is_removable = data.get('is_removable', 'true').lower() == 'true'
            purchase_order_no = data.get('purchase_order_no', '')
            install_time_str = data.get('install_time', '')
            remark = data.get('remark', '')

            with transaction.atomic():
                # 获取静态资产信息
                static_asset = StaticAsset.objects.select_for_update().get(id=static_asset_id)
                
                # 根据静态资产找到或创建对应的Host记录
                parent_host, created = Host.objects.get_or_create(
                    asset_no=static_asset.asset_no,
                    defaults={
                        'hostname': static_asset.asset_no or f"host-{static_asset.id}",
                        'ip': static_asset.ip or '0.0.0.0',
                        'device_model': static_asset.device_model,
                        'status': static_asset.status
                    }
                )

                slot_occupied = AssetRelation.objects.select_for_update()\
                    .filter(parent_asset_id=parent_host.id, slot=slot, is_active=True).exists()
                if slot_occupied:
                    return JsonResponse({'success': False, 'error': f'槽位 {slot} 已被占用'})

                # 获取用户输入的显示名称（可以重复）
                component_display_name = data.get('component_display_name', '').strip()
                
                # 自动生成唯一的hostname
                host_identifier = static_asset.asset_no or static_asset.ip or str(static_asset.id)
                hostname = f"direct-{host_identifier}-{component_name.replace(' ', '-').lower()}-{int(timezone.now().timestamp())}"
                
                child_host = Host.objects.create(
                    hostname=hostname,
                    memo=component_display_name or component_name,  # 使用memo字段存储显示名称
                    asset_no=component_asset_no,  # 使用资产编号
                    device_model=component_model,
                    sn=component_sn,
                    ip='0.0.0.0'
                )

                relation = AssetRelation.objects.create(
                    parent_asset=parent_host,
                    child_asset=child_host,
                    slot=slot,
                    is_removable=is_removable,
                    is_active=True
                )

                install_time = timezone.now()
                if install_time_str:
                    try:
                        from datetime import datetime
                        install_time = datetime.strptime(install_time_str, '%Y-%m-%dT%H:%M')
                        from django.utils.timezone import make_aware
                        install_time = make_aware(install_time)
                    except:
                        pass

                InstallHistory.objects.create(
                    asset_relation=relation,
                    parent_asset=parent_host,
                    child_asset=child_host,
                    install_time=install_time,
                    operator=request.user,
                    operation_type='direct_install',
                    source_type='direct_purchase',
                    purchase_order_no=purchase_order_no or None,
                    remark=remark or f'直接采购安装'
                )

                LifecycleEvent.objects.create(
                    asset=child_host,
                    event_type='purchase',
                    event_time=install_time,
                    operator=request.user,
                    remark=f'直接采购安装到 {parent_host.hostname} 的 {slot} 槽位'
                )

                LifecycleEvent.objects.create(
                    asset=child_host,
                    event_type='deploy',
                    event_time=timezone.now(),
                    operator=request.user,
                    remark=f'安装到主机 {parent_host.hostname} 的 {slot} 槽位'
                )

            return JsonResponse({'success': True, 'relation_id': relation.id})
        except StaticAsset.DoesNotExist:
            return JsonResponse({'success': False, 'error': '静态资产不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_get_relation_detail(request):
    """获取资产关系详情"""
    if request.method == 'GET':
        try:
            relation_id = request.GET.get('relation_id')
            relation = AssetRelation.objects.select_related('child_asset').get(id=relation_id)
            
            data = {
                'id': relation.id,
                'slot': relation.slot,
                'child_asset': {
                    'id': relation.child_asset.id,
                    'hostname': relation.child_asset.hostname,
                    'asset_no': relation.child_asset.asset_no,
                    'name': relation.child_asset.memo,  # 返回memo字段作为显示名称
                    'brand': relation.child_asset.brand,
                    'device_model': relation.child_asset.device_model,
                    'sn': relation.child_asset.sn,
                    'images': relation.child_asset.get_images_list()
                }
            }
            return JsonResponse({'success': True, 'data': data})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持GET请求'})


@login_required
def api_edit_relation(request):
    """编辑资产关系"""
    if request.method == 'POST':
        try:
            data = request.POST
            relation_id = data.get('relation_id')
            asset_no = data.get('asset_no', '')
            name = data.get('name', '')  # 修改为显示名称
            brand = data.get('brand', '')
            device_model = data.get('device_model', '')
            sn = data.get('sn', '')
            slot = data.get('slot', '')

            with transaction.atomic():
                relation = AssetRelation.objects.select_for_update().get(id=relation_id)
                
                # 记录变更信息用于历史记录
                changes = []
                
                # 更新子资产信息
                if asset_no and asset_no != relation.child_asset.asset_no:
                    changes.append(f"资产编号: {relation.child_asset.asset_no or '-'} -> {asset_no}")
                    relation.child_asset.asset_no = asset_no
                if name and name != relation.child_asset.memo:
                    changes.append(f"显示名称: {relation.child_asset.memo or '-'} -> {name}")
                    relation.child_asset.memo = name  # 使用memo字段存储显示名称
                if brand and brand != relation.child_asset.brand:
                    changes.append(f"品牌: {relation.child_asset.brand or '-'} -> {brand}")
                    relation.child_asset.brand = brand
                if device_model and device_model != relation.child_asset.device_model:
                    changes.append(f"型号: {relation.child_asset.device_model or '-'} -> {device_model}")
                    relation.child_asset.device_model = device_model
                if sn and sn != relation.child_asset.sn:
                    changes.append(f"序列号: {relation.child_asset.sn or '-'} -> {sn}")
                    relation.child_asset.sn = sn
                
                # 处理图片上传 - 按年月日创建目录
                # 获取剩余的现有图片路径
                remaining_images = []
                if request.POST.get('remaining_images'):
                    try:
                        remaining_images = json.loads(request.POST.get('remaining_images'))
                    except:
                        pass
                
                # 处理新上传的图片
                new_images = []
                if request.FILES:
                    import datetime
                    date_dir = datetime.datetime.now().strftime('%Y%m%d')
                    for key in request.FILES:
                        file = request.FILES[key]
                        import uuid
                        ext = file.name.split('.')[-1]
                        filename = f"{uuid.uuid4().hex}.{ext}"
                        filepath = os.path.join(settings.MEDIA_ROOT, 'spareparts', date_dir, filename)
                        os.makedirs(os.path.dirname(filepath), exist_ok=True)
                        with open(filepath, 'wb') as f:
                            for chunk in file.chunks():
                                f.write(chunk)
                        new_images.append(f'/media/spareparts/{date_dir}/{filename}')
                
                # 合并剩余图片和新上传的图片
                all_images = remaining_images + new_images
                if all_images:
                    relation.child_asset.images = json.dumps(all_images)
                    if new_images:
                        changes.append(f"添加图片: {len(new_images)} 张")
                else:
                    relation.child_asset.images = ''
                
                relation.child_asset.save()
                
                # 更新槽位信息
                if slot and slot != relation.slot:
                    # 检查新槽位是否被占用
                    slot_occupied = AssetRelation.objects.select_for_update()\
                        .filter(parent_asset_id=relation.parent_asset_id, slot=slot, is_active=True)\
                        .exclude(id=relation_id).exists()
                    if slot_occupied:
                        return JsonResponse({'success': False, 'error': f'槽位 {slot} 已被占用'})
                    
                    changes.append(f"槽位: {relation.slot or '-'} -> {slot}")
                    relation.slot = slot
                    relation.save()
                elif slot:
                    if slot != relation.slot:
                        changes.append(f"槽位: {relation.slot or '-'} -> {slot}")
                    relation.slot = slot
                    relation.save()
                else:
                    relation.save()
                
                # 如果有任何变更，记录历史
                if changes:
                    LifecycleEvent.objects.create(
                        asset=relation.child_asset,
                        event_type='maintenance',
                        event_time=timezone.now(),
                        operator=request.user,
                        remark="编辑资产关系: " + "; ".join(changes)
                    )

            return JsonResponse({'success': True})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


# ==================== 资产关系管理 ====================

@login_required
def api_remove_relation(request):
    """拆卸组件 - 标记资产关系为非活跃"""
    if request.method == 'POST':
        try:
            data = request.POST
            relation_id = data.get('relation_id')
            action = data.get('action', 'uninstall')  # uninstall: 拆卸, scrap: 报废

            with transaction.atomic():
                relation = AssetRelation.objects.select_for_update().get(id=relation_id)
                
                # 标记为非活跃
                relation.is_active = False
                relation.save()
                
                # 添加生命周期事件
                event_type = 'uninstall' if action == 'uninstall' else 'scrap'
                remark = f"从主机 {relation.parent_asset.asset_no} 拆卸" if action == 'uninstall' else f"报废"
                
                LifecycleEvent.objects.create(
                    asset=relation.child_asset,
                    event_type=event_type,
                    event_time=timezone.now(),
                    operator=request.user,
                    remark=remark
                )

            return JsonResponse({'success': True})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_get_relation_history(request):
    """获取资产关系的安装历史"""
    if request.method == 'GET':
        try:
            relation_id = request.GET.get('relation_id')
            relation = AssetRelation.objects.select_related('child_asset').get(id=relation_id)
            
            # 获取子资产的所有生命周期事件
            events = LifecycleEvent.objects.filter(asset_id=relation.child_asset.id)\
                .select_related('operator').order_by('-event_time')
            
            result = []
            for event in events:
                result.append({
                    'id': event.id,
                    'event_type': event.get_event_type_display(),
                    'event_time': event.event_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'operator': event.operator.username if event.operator else '',
                    'remark': event.remark
                })
            
            return JsonResponse({'success': True, 'data': result})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持GET请求'})


@login_required
def api_get_asset_lifecycle(request):
    """获取资产的生命周期历史记录"""
    if request.method == 'GET':
        try:
            asset_id = request.GET.get('asset_id')
            
            # 获取资产的所有生命周期事件
            events = LifecycleEvent.objects.filter(asset_id=asset_id)\
                .select_related('operator').order_by('-event_time')
            
            result = []
            for event in events:
                result.append({
                    'id': event.id,
                    'event_type': event.event_type,
                    'event_type_display': event.get_event_type_display(),
                    'event_time': event.event_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'operator': event.operator.username if event.operator else '',
                    'remark': event.remark if event.remark else '',
                    'created_at': event.created_at.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            return JsonResponse({'success': True, 'data': result})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持GET请求'})


# ==================== 生命周期管理 ====================

@login_required
def lifecycle_event_list(request):
    """生命周期事件列表页面"""
    events = LifecycleEvent.objects.select_related('asset', 'operator')\
        .order_by('-event_time')
    static_assets = StaticAsset.objects.all().order_by('asset_no')
    return render(request, 'cmdb/lifecycle/list.html', {
        'events': events,
        'static_assets': static_assets
    })


@login_required
def api_get_lifecycle_events(request, host_id):
    """获取资产的生命周期事件"""
    events = LifecycleEvent.objects.filter(asset_id=host_id)\
        .select_related('operator').order_by('-event_time')
    result = []
    for event in events:
        result.append({
            'id': event.id,
            'event_type': event.get_event_type_display(),
            'event_time': event.event_time.strftime('%Y-%m-%d %H:%M:%S'),
            'operator': event.operator.username if event.operator else '',
            'remark': event.remark
        })
    return JsonResponse({'events': result})


@login_required
def uninstalled_hardware_list(request):
    """已拆卸硬件列表页面"""
    search_query = request.GET.get('search', '')
    
    # 获取所有已拆卸但未退库的资产关系
    uninstalled_relations = AssetRelation.objects.filter(is_active=False, is_returned=False)\
        .select_related('parent_asset', 'child_asset')\
    
    # 如果有搜索条件，进行过滤
    if search_query:
        uninstalled_relations = uninstalled_relations.filter(
            Q(parent_asset__asset_no__icontains=search_query) |
            Q(child_asset__asset_no__icontains=search_query) |
            Q(child_asset__hostname__icontains=search_query) |
            Q(child_asset__memo__icontains=search_query) |
            Q(child_asset__device_model__icontains=search_query) |
            Q(child_asset__sn__icontains=search_query)
        )
    
    uninstalled_relations = uninstalled_relations.order_by('-updated_at')
    
    return render(request, 'cmdb/asset_relation/uninstalled_list.html', {
        'relations': uninstalled_relations,
        'search_query': search_query
    })


@login_required
def api_add_to_spareparts(request):
    """将已拆卸的硬件添加到备件库"""
    if request.method == 'POST':
        try:
            data = request.POST
            relation_id = data.get('relation_id')
            status = data.get('status', 'in_stock')
            
            # 状态映射：前端值 -> 模型值
            status_mapping = {
                'in_stock': 'in_stock',    # 正常库存 -> 库存中
                'defective': 'maintenance', # 待维修 -> 维修中
                'scrap': 'scrapped'        # 报废 -> 已报废
            }
            # 使用映射后的值，如果没有映射则使用默认值
            mapped_status = status_mapping.get(status, 'in_stock')
            
            with transaction.atomic():
                relation = AssetRelation.objects.select_related('child_asset').get(id=relation_id)
                child_asset = relation.child_asset
                
                # 检查是否已经添加到备件库
                if relation.is_returned:
                    return JsonResponse({'success': False, 'error': '该设备已经添加到备件库，请勿重复添加'})
                
                # 创建备件记录
                sparepart = SparePart.objects.create(
                    asset_code=child_asset.asset_no or '',
                    name=child_asset.memo or child_asset.hostname,
                    brand='',
                    model=child_asset.device_model or '',
                    serial_number=child_asset.sn or '',
                    size='',
                    status=mapped_status,
                    location=relation.slot or '',
                    images=child_asset.images or '',
                    is_installed=False,
                    installed_host_id=None,
                    installed_slot='',
                    purchase_date=None
                )
                
                # 标记为已退库（添加到备件库）
                relation.is_returned = True
                relation.returned_at = timezone.now()
                relation.save()
                
                # 添加生命周期事件
                LifecycleEvent.objects.create(
                    asset=child_asset,
                    event_type='maintenance',
                    event_time=timezone.now(),
                    operator=request.user,
                    remark=f'已添加到备件库，状态: {sparepart.get_status_display()}'
                )
                
                log_operation(request.user, 'create', f'备件: {sparepart.name}', '从已拆卸硬件添加备件', request.META.get('REMOTE_ADDR'))
            
            return JsonResponse({'success': True, 'message': '已成功添加到备件库'})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def returned_devices_list(request):
    """已退库设备列表页面"""
    search_query = request.GET.get('search', '')
    
    # 获取所有已退库的资产关系
    returned_relations = AssetRelation.objects.filter(is_returned=True)\
        .select_related('parent_asset', 'child_asset')
    
    # 如果有搜索条件，进行过滤
    if search_query:
        returned_relations = returned_relations.filter(
            Q(parent_asset__asset_no__icontains=search_query) |
            Q(child_asset__asset_no__icontains=search_query) |
            Q(child_asset__hostname__icontains=search_query) |
            Q(child_asset__memo__icontains=search_query) |
            Q(child_asset__device_model__icontains=search_query) |
            Q(child_asset__sn__icontains=search_query)
        )
    
    returned_relations = returned_relations.order_by('-returned_at')
    
    return render(request, 'cmdb/asset_relation/returned_list.html', {
        'relations': returned_relations,
        'search_query': search_query
    })


@login_required
def api_batch_return_to_warehouse(request):
    """批量退库API"""
    if request.method == 'POST':
        try:
            data = request.POST
            relation_ids = data.getlist('relation_ids[]')
            
            if not relation_ids:
                return JsonResponse({'success': False, 'error': '请选择要退库的设备'})
            
            with transaction.atomic():
                for relation_id in relation_ids:
                    relation = AssetRelation.objects.get(id=relation_id)
                    
                    # 检查是否已经退库
                    if relation.is_returned:
                        continue
                    
                    # 标记为已退库
                    relation.is_returned = True
                    relation.returned_at = timezone.now()
                    relation.save()
                    
                    # 添加生命周期事件
                    LifecycleEvent.objects.create(
                        asset=relation.child_asset,
                        event_type='maintenance',
                        event_time=timezone.now(),
                        operator=request.user,
                        remark=f'已退库'
                    )
                
                log_operation(request.user, 'update', f'批量退库 {len(relation_ids)} 个设备', '批量退库操作', request.META.get('REMOTE_ADDR'))
            
            return JsonResponse({'success': True, 'message': f'已成功退库 {len(relation_ids)} 个设备'})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_cancel_return(request):
    """取消退库API"""
    if request.method == 'POST':
        try:
            data = request.POST
            relation_id = data.get('relation_id')
            
            with transaction.atomic():
                relation = AssetRelation.objects.get(id=relation_id)
                
                # 检查是否已退库
                if not relation.is_returned:
                    return JsonResponse({'success': False, 'error': '设备尚未退库'})
                
                # 取消退库
                relation.is_returned = False
                relation.returned_at = None
                relation.save()
                
                # 添加生命周期事件
                LifecycleEvent.objects.create(
                    asset=relation.child_asset,
                    event_type='maintenance',
                    event_time=timezone.now(),
                    operator=request.user,
                    remark=f'取消退库'
                )
                
                log_operation(request.user, 'update', f'取消退库: {relation.child_asset.memo or relation.child_asset.hostname}', '取消退库操作', request.META.get('REMOTE_ADDR'))
            
            return JsonResponse({'success': True, 'message': '已成功取消退库，设备已恢复到已拆卸状态'})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_delete_returned_record(request):
    """删除已退库记录API"""
    if request.method == 'POST':
        try:
            data = request.POST
            relation_id = data.get('relation_id')
            
            with transaction.atomic():
                relation = AssetRelation.objects.select_related('child_asset').get(id=relation_id)
                child_asset_name = relation.child_asset.memo or relation.child_asset.hostname
                
                # 删除资产关系记录
                relation.delete()
                
                log_operation(request.user, 'delete', f'删除已退库记录: {child_asset_name}', '删除已退库记录', request.META.get('REMOTE_ADDR'))
            
            return JsonResponse({'success': True, 'message': '已成功删除记录'})
        except AssetRelation.DoesNotExist:
            return JsonResponse({'success': False, 'error': '资产关系不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_add_lifecycle_event(request):
    """添加生命周期事件"""
    if request.method == 'POST':
        try:
            data = request.POST
            static_asset_id = data.get('asset_id')  # 现在是StaticAsset的ID
            event_type = data.get('event_type')
            remark = data.get('remark', '')

            # 获取静态资产信息
            static_asset = StaticAsset.objects.get(id=static_asset_id)
            
            # 根据静态资产找到或创建对应的Host记录
            host, created = Host.objects.get_or_create(
                asset_no=static_asset.asset_no,
                defaults={
                    'hostname': static_asset.asset_no or f"host-{static_asset.id}",
                    'ip': static_asset.ip or '0.0.0.0',
                    'device_model': static_asset.device_model,
                    'status': static_asset.status
                }
            )

            event = LifecycleEvent.objects.create(
                asset=host,
                event_type=event_type,
                event_time=timezone.now(),
                operator=request.user,
                remark=remark
            )
            return JsonResponse({'success': True, 'event_id': event.id})
        except StaticAsset.DoesNotExist:
            return JsonResponse({'success': False, 'error': '静态资产不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_delete_lifecycle_event(request, event_id):
    """删除生命周期事件"""
    if request.method == 'POST':
        try:
            event = LifecycleEvent.objects.get(id=event_id)
            event.delete()
            return JsonResponse({'success': True})
        except LifecycleEvent.DoesNotExist:
            return JsonResponse({'success': False, 'error': '生命周期事件不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def api_batch_delete_lifecycle_events(request):
    """批量删除生命周期事件"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            event_ids = data.get('event_ids', [])
            
            if not event_ids:
                return JsonResponse({'success': False, 'error': '请选择要删除的记录'})
            
            LifecycleEvent.objects.filter(id__in=event_ids).delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


# ==================== 办公机配件管理 ====================

@login_required
def office_parts_list(request):
    """办公机配件列表页面"""
    parts = OfficePart.objects.all().order_by('-created_at')
    
    # 搜索和筛选
    search_keyword = request.GET.get('keyword', '')
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    health_filter = request.GET.get('health', '')
    
    if search_keyword:
        parts = parts.filter(
            Q(name__icontains=search_keyword) |
            Q(brand__icontains=search_keyword) |
            Q(model__icontains=search_keyword) |
            Q(serial_number__icontains=search_keyword) |
            Q(source_computer__icontains=search_keyword)
        )
    
    if category_filter:
        parts = parts.filter(category=category_filter)
    
    if status_filter:
        parts = parts.filter(status=status_filter)
    
    context = {
        'parts': parts,
        'search_keyword': search_keyword,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'categories': OfficePart.CATEGORY_CHOICES,
        'statuses': OfficePart.STATUS_CHOICES,
    }
    return render(request, 'cmdb/office_parts/list.html', context)


@login_required
def office_part_add(request):
    """添加办公机配件"""
    if request.method == 'POST':
        try:
            part = OfficePart()
            part.name = request.POST.get('name')
            part.category = request.POST.get('category', 'other')
            part.brand = request.POST.get('brand', '')
            part.model = request.POST.get('model', '')
            part.serial_number = request.POST.get('serial_number') or None
            part.source_computer = request.POST.get('source_computer', '')
            part.status = request.POST.get('status', 'in_stock')
            part.dismantle_date = request.POST.get('dismantle_date') or None
            part.location = request.POST.get('location', '')
            part.purchase_date = request.POST.get('purchase_date') or None
            part.remark = request.POST.get('remark', '')
            
            # 处理图片上传 - 按年月日创建目录
            images = []
            if request.FILES:
                import datetime
                date_dir = datetime.datetime.now().strftime('%Y%m%d')
                for key in request.FILES:
                    file = request.FILES[key]
                    import uuid
                    ext = file.name.split('.')[-1]
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    filepath = os.path.join(settings.MEDIA_ROOT, 'spareparts', date_dir, filename)
                    os.makedirs(os.path.dirname(filepath), exist_ok=True)
                    with open(filepath, 'wb') as f:
                        for chunk in file.chunks():
                            f.write(chunk)
                    images.append(f'/media/spareparts/{date_dir}/{filename}')
            
            if images:
                part.images = json.dumps(images)
            
            part.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def office_part_edit(request, part_id):
    """编辑办公机配件"""
    if request.method == 'GET':
        try:
            part = OfficePart.objects.get(id=part_id)
            return JsonResponse({
                'success': True,
                'data': {
                    'id': part.id,
                    'name': part.name,
                    'category': part.category,
                    'brand': part.brand,
                    'model': part.model,
                    'serial_number': part.serial_number or '',
                    'source_computer': part.source_computer,
                    'status': part.status,
                    'dismantle_date': part.dismantle_date.strftime('%Y-%m-%d') if part.dismantle_date else '',
                    'location': part.location,
                    'purchase_date': part.purchase_date.strftime('%Y-%m-%d') if part.purchase_date else '',
                    'remark': part.remark,
                    'images': part.get_images_list(),
                }
            })
        except OfficePart.DoesNotExist:
            return JsonResponse({'success': False, 'error': '配件不存在'})
    elif request.method == 'POST':
        try:
            part = OfficePart.objects.get(id=part_id)
            part.name = request.POST.get('name')
            part.category = request.POST.get('category', 'other')
            part.brand = request.POST.get('brand', '')
            part.model = request.POST.get('model', '')
            part.serial_number = request.POST.get('serial_number') or None
            part.source_computer = request.POST.get('source_computer', '')
            part.status = request.POST.get('status', 'in_stock')
            part.dismantle_date = request.POST.get('dismantle_date') or None
            part.location = request.POST.get('location', '')
            part.purchase_date = request.POST.get('purchase_date') or None
            part.remark = request.POST.get('remark', '')
            
            # 处理图片上传 - 追加到现有图片
            if request.FILES:
                # 获取现有图片列表
                existing_images = part.get_images_list()
                new_images = []
                import datetime
                date_dir = datetime.datetime.now().strftime('%Y%m%d')
                for key in request.FILES:
                    file = request.FILES[key]
                    import uuid
                    ext = file.name.split('.')[-1]
                    filename = f"{uuid.uuid4().hex}.{ext}"
                    filepath = os.path.join(settings.MEDIA_ROOT, 'spareparts', date_dir, filename)
                    os.makedirs(os.path.dirname(filepath), exist_ok=True)
                    with open(filepath, 'wb') as f:
                        for chunk in file.chunks():
                            f.write(chunk)
                    new_images.append(f'/media/spareparts/{date_dir}/{filename}')
                
                if new_images:
                    # 追加到现有图片
                    all_images = existing_images + new_images
                    part.images = json.dumps(all_images)
            
            part.save()
            return JsonResponse({'success': True})
        except OfficePart.DoesNotExist:
            return JsonResponse({'success': False, 'error': '配件不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持GET和POST请求'})


@login_required
def office_part_delete(request, part_id):
    """删除办公机配件"""
    if request.method == 'POST':
        try:
            part = OfficePart.objects.get(id=part_id)
            part.delete()
            return JsonResponse({'success': True})
        except OfficePart.DoesNotExist:
            return JsonResponse({'success': False, 'error': '配件不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def office_part_update_status(request):
    """更新配件状态"""
    if request.method == 'POST':
        try:
            part_id = request.POST.get('part_id')
            new_status = request.POST.get('status')
            
            part = OfficePart.objects.get(id=part_id)
            part.status = new_status
            part.save()
            return JsonResponse({'success': True})
        except OfficePart.DoesNotExist:
            return JsonResponse({'success': False, 'error': '配件不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


@login_required
def office_part_batch_delete(request):
    """批量删除办公机配件"""
    if request.method == 'POST':
        try:
            import json
            ids = json.loads(request.POST.get('ids', '[]'))
            
            if not ids:
                return JsonResponse({'success': False, 'error': '请选择要删除的配件'})
            
            OfficePart.objects.filter(id__in=ids).delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': '只支持POST请求'})


# ==================== 组件历史查询 ====================

@login_required
def component_history(request, host_id):
    """组件历史去向查询"""
    host = get_object_or_404(Host, id=host_id)
    
    # 查询该组件作为子资产的所有历史关系
    relations = AssetRelation.objects.filter(child_asset=host)\
        .select_related('parent_asset')
    
    # 获取所有安装历史
    history_data = []
    for rel in relations:
        histories = InstallHistory.objects.filter(asset_relation=rel)\
            .select_related('operator').order_by('install_time')
        for h in histories:
            history_data.append({
                'parent_asset': rel.parent_asset,
                'slot': rel.slot,
                'install_time': h.install_time,
                'uninstall_time': h.uninstall_time,
                'operator': h.operator,
                'operation_type': h.operation_type,
                'remark': h.remark
            })
    
    history_data.sort(key=lambda x: x['install_time'], reverse=True)
    
    return render(request, 'cmdb/component_history.html', {
        'component': host,
        'history': history_data
    })
